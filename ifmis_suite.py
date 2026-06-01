# =========================================================
# IFMIS SUITE - INTEGRATED AUTOMATION PORTAL (V1.0)
# Combines Voucher (v4.1), Paybill (v3.8), and POL Downloading (PLM7)
# REQUIRED: pip install openpyxl playwright pywin32 customtkinter pdfplumber pandas tkcalendar pillow pytesseract
# =========================================================

import os
os.environ["NO_PROXY"] = "localhost,127.0.0.1"
import sys
import re
import time
import base64
import calendar
import threading
import asyncio
import datetime
import traceback
import zipfile
import json
import socket
import subprocess
import urllib.request
import urllib.parse
import platform
import getpass
import uuid
import tkinter as tk
from tkinter import messagebox, filedialog
import customtkinter as ctk

# Placeholders for deferred imports to speed up startup splash screen
load_workbook = None
pdfplumber = None
sync_playwright = None
PlaywrightTimeoutError = None
pytesseract = None
Image = None
ImageEnhance = None
ImageFilter = None
io = None
Calendar = None
pd = None

TESSERACT_AVAILABLE = False
TKCALENDAR_AVAILABLE = False
PANDAS_AVAILABLE = False

def split_date_range_by_months(from_date_str, to_date_str):
    try:
        fd, fm, fy = map(int, from_date_str.split('/'))
        td, tm, ty = map(int, to_date_str.split('/'))
        start_date = datetime.date(fy, fm, fd)
        end_date = datetime.date(ty, tm, td)
    except Exception:
        # Fallback if parsing fails - just return a single block containing the input strings
        return [(from_date_str, to_date_str)]
        
    if start_date > end_date:
        return []
        
    intervals = []
    current_start = start_date
    while current_start <= end_date:
        # End of current calendar month
        last_day_of_month = calendar.monthrange(current_start.year, current_start.month)[1]
        month_end_date = datetime.date(current_start.year, current_start.month, last_day_of_month)
        
        current_end = min(month_end_date, end_date)
        
        intervals.append((
            current_start.strftime("%d/%m/%Y"),
            current_end.strftime("%d/%m/%Y")
        ))
        
        # Advance to the first day of next month
        current_start = current_end + datetime.timedelta(days=1)
        
    return intervals

def load_engines(progress_callback=None):
    global load_workbook, pdfplumber, sync_playwright, PlaywrightTimeoutError
    global pytesseract, Image, ImageEnhance, ImageFilter, io, Calendar, pd
    global TESSERACT_AVAILABLE, TKCALENDAR_AVAILABLE, PANDAS_AVAILABLE

    if progress_callback: progress_callback("Loading Excel utilities...", 55)
    from openpyxl import load_workbook
    
    if progress_callback: progress_callback("Loading PDF Plumber extraction...", 65)
    import pdfplumber
    
    if progress_callback: progress_callback("Loading Playwright automation...", 75)
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError
    PlaywrightTimeoutError = PWTimeoutError
    
    if progress_callback: progress_callback("Loading OCR CAPTCHA engines...", 85)
    try:
        import pytesseract
        from PIL import Image as PILImage, ImageEnhance as PILImageEnhance, ImageFilter as PILImageFilter
        import io as io_module
        Image = PILImage
        ImageEnhance = PILImageEnhance
        ImageFilter = PILImageFilter
        io = io_module
        TESSERACT_AVAILABLE = True
        
        if sys.platform == "win32":
            possible_paths = [
                r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                r"C:\Users\{}\AppData\Local\Programs\Tesseract-OCR\tesseract.exe".format(os.getenv('USERNAME')),
                r"C:\Tesseract-OCR\tesseract.exe",
            ]
            for path in possible_paths:
                if os.path.exists(path):
                    pytesseract.pytesseract.tesseract_cmd = path
                    break
            else:
                import shutil
                tess_in_path = shutil.which("tesseract") or shutil.which("tesseract.exe")
                if tess_in_path:
                    pytesseract.pytesseract.tesseract_cmd = tess_in_path
    except ImportError:
        TESSERACT_AVAILABLE = False

    if progress_callback: progress_callback("Loading calendar picker...", 90)
    try:
        from tkcalendar import Calendar as CalendarClass
        Calendar = CalendarClass
        TKCALENDAR_AVAILABLE = True
    except ImportError:
        TKCALENDAR_AVAILABLE = False

    if progress_callback: progress_callback("Loading data analysis modules...", 95)
    try:
        import pandas as pandas_module
        pd = pandas_module
        PANDAS_AVAILABLE = True
        # Explicitly import xlrd and lxml so PyInstaller packages them as hidden dependencies for Excel/HTML readers
        import xlrd
        import lxml
    except ImportError:
        PANDAS_AVAILABLE = False

    if progress_callback: progress_callback("Engines loaded successfully!", 100)

# Windows-specific libraries
if sys.platform == "win32":
    import win32gui
    import win32con
    import winreg

# --- UI STYLE CONFIGURATION ---
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

TOOL_VERSION = "IFMIS Suite v1.0"
RUN_ID = str(uuid.uuid4())[:8]

# Direct Multi-Column Google Sheet Logging (Apps Script Web App URL)
# Paste your deployed Google Apps Script Web App URL here, e.g. "https://script.google.com/macros/s/XXXXX/exec"
TELEMETRY_URL = "https://script.google.com/macros/s/AKfycbz3bZ6vhzPIZS8Wz3K3SGtyg-qAiHqfstkh4AJDQ2sQzCxlPidjKh9pe9eNcjJZGCGP/exec"

# =========================================================
# SYSTEM DETECTIONS & EXPORT SUPPORT
# =========================================================
def get_system_info():
    try:
        return {
            "os": platform.platform(),
            "os_version": platform.version(),
            "architecture": platform.machine(),
            "windows_user": getpass.getuser(),
            "python_version": platform.python_version()
        }
    except Exception:
        return {}

def get_playwright_version():
    try:
        import pkg_resources
        return pkg_resources.get_distribution("playwright").version
    except Exception:
        try:
            from importlib.metadata import version as pkg_version
            return pkg_version("playwright")
        except Exception:
            return "UNKNOWN"

def get_machine_id():
    try:
        output = subprocess.check_output('wmic csproduct get uuid', creationflags=0x08000000).decode('utf-8').strip()
        lines = [line.strip() for line in output.split('\n') if line.strip() != '']
        if len(lines) > 1 and len(lines[1]) > 10: return lines[1]
    except Exception: pass

    try:
        cmd = ['powershell', '-NoProfile', '-Command', '(Get-CimInstance -Class Win32_ComputerSystemProduct).UUID']
        output = subprocess.check_output(cmd, creationflags=0x08000000).decode('utf-8').strip()
        if output and len(output) > 10: return output
    except Exception: pass

    try:
        registry = winreg.ConnectRegistry(None, winreg.HKEY_LOCAL_MACHINE)
        key = winreg.OpenKey(registry, r"SOFTWARE\Microsoft\Cryptography")
        machine_guid, _ = winreg.QueryValueEx(key, "MachineGuid")
        if machine_guid: return machine_guid
    except Exception: pass

    return "UNKNOWN_MACHINE"

def get_ip_address():
    local_ip = "UNKNOWN_LOCAL"
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
    except Exception: pass

    public_ip = "BLOCKED_BY_FIREWALL"
    ip_apis = ["http://checkip.amazonaws.com", "https://icanhazip.com", "https://api.ipify.org"]
    for api in ip_apis:
        try:
            req = urllib.request.Request(api, headers={'User-Agent': 'Mozilla/5.0'})
            response = urllib.request.urlopen(req, timeout=4).read().decode('utf-8').strip()
            if re.match(r"^\d{1,3}(\.\d{1,3}){3}$", response):
                public_ip = response
                break
        except Exception: continue

    return f"Pub:{public_ip} | Loc:{local_ip}"

def send_silent_log(username, combined_data):
    try:
        # If a custom Apps Script URL is configured, POST the JSON payload directly for separate cells
        if TELEMETRY_URL and TELEMETRY_URL.strip():
            req = urllib.request.Request(
                TELEMETRY_URL.strip(),
                data=combined_data.encode("utf-8"),
                headers={"Content-Type": "application/json"}
            )
            urllib.request.urlopen(req, timeout=5)
            return

        # Fallback to standard Google Form response
        google_form_url = "https://docs.google.com/forms/d/e/1FAIpQLSeStAIOE96quY9zPR4GkScQf1ZHAuoKfCHnas-37MEDHDJzkg/formResponse"
        form_data = {"entry.872389760": username, "entry.968476556": combined_data}
        data = urllib.parse.urlencode(form_data).encode("utf-8")
        req = urllib.request.Request(google_form_url, data=data)
        urllib.request.urlopen(req, timeout=3)
    except Exception: pass

def check_security(operator_name):
    my_hwid = get_machine_id()
    my_ip = get_ip_address()
    try: pc_name = socket.gethostname()
    except Exception: pc_name = "UNKNOWN_PC"
    
    # 1. Pastebin HWID and global block checks (from 3.6 - final)
    try:
        kill_switch_url = "https://pastebin.com/raw/sUxsMvyJ"
        req = urllib.request.Request(kill_switch_url, headers={'User-Agent': 'Mozilla/5.0'})
        response = urllib.request.urlopen(req, timeout=5).read().decode('utf-8')
        config = json.loads(response)
        if not config.get("global_enabled", True):
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Access Denied", "The IFMIS Suite has been globally disabled by Nitin Arya.")
            sys.exit()
        if my_hwid in config.get("blocked_hardware_ids", []):
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("License Revoked", "Security Violation: This machine has been blocked.")
            sys.exit()
    except Exception:
        # Fall through gracefully if pastebin is unreachable (e.g. offline dry run)
        pass

    # 2. Expiration Date Check (Valid until end of 2050)
    expiration_date = datetime.date(2050, 12, 31)
    if datetime.date.today() > expiration_date:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("License Expired", "This version of the IFMIS Suite has expired.")
        sys.exit()

    # 3. Startup Telemetry logging (from 4.1 Excel Mode)
    system_info = get_system_info()
    telemetry_payload = {
        "run_id": RUN_ID,
        "event": "startup",
        "operator": operator_name,
        "pc_name": pc_name,
        "hwid": my_hwid,
        "ip": my_ip,
        "tool_version": TOOL_VERSION,
        "os": system_info.get("os"),
        "os_version": system_info.get("os_version"),
        "architecture": system_info.get("architecture"),
        "windows_user": system_info.get("windows_user"),
        "python_version": system_info.get("python_version"),
        "playwright_version": get_playwright_version(),
        "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }

    threading.Thread(
        target=send_silent_log,
        args=(operator_name, json.dumps(telemetry_payload)),
        daemon=True
    ).start()

def load_major_heads():
    mh_set = set()
    possible_paths = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "MP_Major_Heads.csv"),
        os.path.join(os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__)), "MP_Major_Heads.csv")
    ]
    for path in possible_paths:
        try:
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    for line in f:
                        parts = re.findall(r"\d{4}", line)
                        for p in parts: mh_set.add(p.strip())
                break
        except Exception: pass
    return mh_set

KNOWN_MAJOR_HEADS = load_major_heads()

# Determine paths
if getattr(sys, "frozen", False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

PW_BROWSER_DIR = os.path.join(APP_DIR, "pw-browsers")

if os.path.isdir(PW_BROWSER_DIR):
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = PW_BROWSER_DIR

print(f"Using Playwright browsers: {os.environ.get('PLAYWRIGHT_BROWSERS_PATH')}")


# =========================================================
# TREASURY DATABASE & DDO EXTRACTOR
# =========================================================
TREASURY_IDS = {
    "010-BAL-BALAGHAT TREASURY": "18312~010~BALAGHAT TREASURY",
    "020-BAD-Barwani District Treasury": "18071~020~Barwani District Treasury",
    "030-BET-Betul District Treasury": "17863~030~Betul District Treasury",
    "040-BHI-Bhind Treasury": "17575~040~Bhind Treasury",
    "050-BPL-District Treasury Office Bhopal": "17472~050~District Treasury Office Bhopal",
    "051-VIN-Vindhyachal Treasury": "17465~051~Vindhyachal Treasury",
    "052-VAL-Vallabh Bhawan Treasury, Bhopal": "17470~052~Vallabh Bhawan Treasury, Bhopal",
    "054-CTB-Principal Treasury Officer Centre Treasury Bhopal": "50027814~054~Principal Treasury Officer Centre Treasury Bhopal",
    "060-CHA-Chhatarpur Treasury": "17081~060~Chhatarpur Treasury",
    "070-CHI-Chhindwara Treasury": "16858~070~Chhindwara Treasury",
    "080-DAM-Damoh Treasury": "16573~080~Damoh Treasury",
    "090-DAT-Datia Treasury": "16419~090~Datia Treasury",
    "100-DEW-Dewas Treasury": "16304~100~Dewas Treasury",
    "110-DHA-Dhar Treasury": "16117~110~Dhar Treasury",
    "120-DIN-Dindori Treasury": "15743~120~Dindori Treasury",
    "130-GUN-Guna Treasury": "15710~130~Guna Treasury",
    "140-GWL-Gwalior Gorkhi Treasury": "15549~140~Gwalior Gorkhi Treasury",
    "141-MML-Motimahal Gwalior Treasury": "15547~141~Motimahal Gwalior Treasury",
    "150-HAR-Harda Treasury": "15292~150~Harda Treasury",
    "160-HOS-Hoshangabad Treasury": "15182~160~Hoshangabad Treasury",
    "170-IND-Indore District Treasury": "14925~170~Indore District Treasury",
    "171-INC-Indore City Treasury": "14929~171~Indore City Treasury",
    "180-JBP-Jabalpur District Treasury": "14605~180~Jabalpur District Treasury",
    "181-JBP-Jabalpur City Treasury": "14593~181~Jabalpur City Treasury",
    "190-JHA-Jhabua Treasury": "14278~190~Jhabua Treasury",
    "200-KAT-Katni Treasury": "14080~200~Katni Treasury",
    "210-KHA-Khandva Treasury": "13880~210~Khandva Treasury",
    "220-KAR-Khargaun Treasury": "13688~220~Khargaun Treasury",
    "230-MAN-Mandla Treasury": "13412~230~Mandla Treasury",
    "240-MND-Mandsaur Treasury": "13136~240~Mandsaur Treasury",
    "250-MOE-Morena Treasury": "12959~250~Morena Treasury",
    "260-NAR-Narsinghpur Treasury": "12775~260~Narsinghpur Treasury",
    "270-NEE-Neemuch District Treasury": "12605~270~Neemuch District Treasury",
    "280-PAN-Panna Treasury": "12473~280~Panna Treasury",
    "290-RIS-Raisen Treasury": "12326~290~Raisen Treasury",
    "300-RAJ-Rajgarh Treasury": "12147~300~Rajgarh Treasury",
    "310-RAT-Ratlam Treasury": "11969~310~Ratlam Treasury",
    "320-REW-Rewa Treasury": "11741~320~Rewa Treasury",
    "330-SAG-Sagar Treasury": "11416~330~Sagar Treasury",
    "340-SAT-Satna Treasury": "11125~340~Satna Treasury",
    "350-SEH-Sehore Treasury": "10878~350~Sehore Treasury",
    "360-SEO-Seoni Treasury": "10703~360~Seoni Treasury",
    "370-SHA-Shahdol Treasury": "10460~370~Shahdol Treasury",
    "380-SAJ-Shajapur Treasury": "10266~380~Shajapur Treasury",
    "390-SHI-Shivpuri Treasury": "10120~390~Shivpuri Treasury",
    "400-SHE-Sheopur Treasury": "9963~400~Sheopur Treasury",
    "410-SID-Seedhi Treasury": "9855~410~Seedhi Treasury",
    "420-TIK-Tikamgarh Treasury": "9660~420~Tikamgarh Treasury",
    "430-UJJ-Ujjain Treasury": "9491~430~Ujjain Treasury",
    "440-UMA-Umaria Treasury": "9242~440~Umaria Treasury",
    "450-VID-Vidisha Treasury": "9123~450~Vidisha Treasury",
    "460-ANU-Anupur Treasury": "18665~460~Anupur Treasury",
    "470-ASH-Ashoknagar Treasury": "18483~470~Ashoknagar Treasury",
    "480-BUR-Burhanpur Treasury": "18393~480~Burhanpur Treasury",
    "490-ALI-Alirajpur Treasury": "8950~490~Alirajpur Treasury",
    "500-SNG-Singrauli Treasury": "8833~500~Singrauli Treasury",
    "510-AGR-Agar Malwa Treasury": "18736~510~Agar Malwa Treasury",
    "520-NIW-District Treasury Office Niwari": "50027813~520~District Treasury Office Niwari"
}

def parse_ddo_file():
    mapping = {}
    try:
        if getattr(sys, 'frozen', False):
            base_path = getattr(sys, '_MEIPASS', os.path.dirname(sys.executable))
        else:
            base_path = os.path.dirname(os.path.abspath(__file__))
        
        ddo_path = os.path.join(base_path, "ddo.txt")
        if not os.path.exists(ddo_path) and getattr(sys, 'frozen', False):
            ddo_path = os.path.join(os.path.dirname(sys.executable), "ddo.txt")
        if os.path.exists(ddo_path):
            with open(ddo_path, "r", encoding="utf-8") as f:
                content = f.read()
            blocks = re.split(r"=== (.*?) ===", content)
            for i in range(1, len(blocks), 2):
                treasury_name = blocks[i].strip().replace("(V-)", "")
                ddos = [line.strip() for line in blocks[i + 1].split("\n") if line.strip()]
                mapping[treasury_name] = ddos
    except Exception as e:
        print(f"Error loading ddo.txt: {e}")
    return mapping

DDO_DATABASE = parse_ddo_file()


# =========================================================
# EXCEPTIONS & DIALOGS
# =========================================================
class SessionExpiredError(Exception):
    pass

class DatePickerModal(ctk.CTkToplevel):
    def __init__(self, parent, target_entry):
        super().__init__(parent)
        self.title("Select Date")
        self.geometry("310x340")
        self.configure(fg_color="#0d1117")
        self.transient(parent)
        self.grab_set()
        self.target_entry = target_entry

        if TKCALENDAR_AVAILABLE:
            self.cal = Calendar(self, selectmode='day', date_pattern='dd/mm/yyyy',
                           background='#0d1117', foreground='white', borderwidth=0,
                           headersbackground='#161b22', headersforeground='#00c9a7',
                           normalbackground='#21262d', normalforeground='white',
                           bottombackground='#161b22', selectbackground='#00c9a7')
            self.cal.pack(fill="both", expand=True, padx=10, pady=10)
            btn = ctk.CTkButton(
                self, text="Confirm Date", command=self.set_date, 
                fg_color="#00c9a7", hover_color="#00e5bf", text_color="#0d1117",
                font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), corner_radius=8
            )
            btn.pack(pady=(0, 15))
        else:
            lbl = ctk.CTkLabel(
                self, text="tkcalendar library not found.\nPlease enter date manually in DD/MM/YYYY format.", 
                font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#f8fafc", wraplength=260
            )
            lbl.pack(pady=30, padx=20)
            btn = ctk.CTkButton(
                self, text="Close", command=self.destroy,
                fg_color="#3b82f6", hover_color="#1d4ed8", text_color="#f8fafc",
                font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), corner_radius=8
            )
            btn.pack(pady=10)

    def set_date(self):
        self.target_entry.delete(0, "end")
        self.target_entry.insert(0, self.cal.get_date())
        self.destroy()

class TreasuryPickerModal(ctk.CTkToplevel):
    def __init__(self, parent, current_treasury, on_select_callback):
        super().__init__(parent)
        self.title("Select Treasury")
        self.geometry("450x550")
        self.configure(fg_color="#0d1117")
        self.transient(parent)
        self.grab_set()
        self.on_select = on_select_callback

        self.search_var = ctk.StringVar()
        self.search_entry = ctk.CTkEntry(
            self, textvariable=self.search_var, placeholder_text="🔍 Search Treasury by Name or Code...", 
            fg_color="#0d1117", border_color="#30363d", text_color="#f8fafc",
            font=ctk.CTkFont(family="Segoe UI", size=11), corner_radius=8
        )
        self.search_entry.pack(fill="x", padx=15, pady=15)
        self.search_entry.bind("<KeyRelease>", self.filter_treasuries)

        self.scroll_frame = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scroll_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))

        self.buttons = []
        self.populate()

    def populate(self, filter_text=""):
        for btn in self.buttons:
            btn.destroy()
        self.buttons.clear()

        for t_name in sorted(TREASURY_IDS.keys()):
            if filter_text.lower() in t_name.lower():
                btn = ctk.CTkButton(
                    self.scroll_frame, 
                    text=t_name, 
                    anchor="w", 
                    fg_color="transparent",
                    hover_color="#0c362d", 
                    text_color="#f8fafc", 
                    font=ctk.CTkFont(family="Segoe UI", size=11),
                    corner_radius=6,
                    command=lambda t=t_name: self.select(t)
                )
                btn.pack(fill="x", pady=2)
                self.buttons.append(btn)

    def filter_treasuries(self, event=None):
        self.populate(self.search_var.get())

    def select(self, t_name):
        self.on_select(t_name)
        self.destroy()


class OperatorLoginDialog(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("IFMIS Suite - Login")
        self.geometry("360x200")
        self.configure(fg_color="#0d1117")
        self.resizable(False, False)
        
        # Center window
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width - 360) // 2
        y = (screen_height - 200) // 2
        self.geometry(f"360x200+{x}+{y}")
        self.attributes("-topmost", True)
        self.grab_set()
        
        self.result = None
        
        ctk.CTkLabel(
            self, text="👤 Operator Authentication", 
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"), text_color="#00c9a7"
        ).pack(pady=(20, 10))
        
        self.entry = ctk.CTkEntry(
            self, placeholder_text="Enter Operator Name...", width=280, height=32,
            fg_color="#0d1117", border_color="#30363d", text_color="#f8fafc",
            font=ctk.CTkFont(family="Segoe UI", size=11), corner_radius=8
        )
        self.entry.pack(pady=10)
        self.entry.focus_set()
        
        btn = ctk.CTkButton(
            self, text="Confirm & Login", height=32, width=140,
            fg_color="#00c9a7", hover_color="#00e5bf", text_color="#0d1117",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), corner_radius=8,
            command=self.on_submit
        )
        btn.pack(pady=15)
        
        self.entry.bind("<Return>", lambda e: self.on_submit())
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        
    def on_submit(self):
        name = self.entry.get().strip()
        if name:
            self.result = "".join(c for c in name if c.isalnum() or c in (" ", "_", "-")).strip()
            self.destroy()
            
    def on_close(self):
        self.result = None
        self.destroy()


# =========================================================
# MAIN SUITE APPLICATION PORTAL
# =========================================================
class IFMISSuiteApp(ctk.CTk):

    TREASURY_MAP = {
        "BAL": "010", "BAD": "020", "BET": "030", "BHI": "040", "BPL": "050",
        "VIN": "051", "VAL": "052", "CTB": "054", "CHA": "060", "CHI": "070",
        "DAM": "080", "DAT": "090", "DEW": "100", "DHA": "110", "DIN": "120",
        "GUN": "130", "GWL": "140", "MML": "141", "HAR": "150", "HOS": "160",
        "IND": "170", "INC": "171", "JBP": "180", "JBC": "181", "JHA": "190",
        "KAT": "200", "KHA": "210", "KAR": "220", "MAN": "230", "MND": "240",
        "MOR": "250", "NAR": "260", "NEE": "270", "PAN": "280", "RIS": "290",
        "RAJ": "300", "RAT": "310", "REW": "320", "SAG": "330", "SAT": "340",
        "SEH": "350", "SEO": "360", "SHA": "370", "SAJ": "380", "SHI": "390",
        "SHE": "400", "SID": "410", "TIK": "420", "UJJ": "430", "UMA": "440",
        "VID": "450", "ANU": "460", "ASH": "470", "BUR": "480", "ALI": "490",
        "SNG": "500", "AGR": "510", "NIW": "520",
    }

    def __init__(self, operator_name):
        super().__init__()

        # Prevent Windows system and screen from going to sleep while IFMIS Suite is open
        if sys.platform == "win32":
            try:
                import ctypes
                ctypes.windll.kernel32.SetThreadExecutionState(0x80000000 | 0x00000001 | 0x00000002)
            except Exception: pass

        # Config Variables
        self.var_stealth = ctk.BooleanVar(value=False)
        self.var_archive = ctk.BooleanVar(value=False)
        self.var_dryrun = ctk.BooleanVar(value=False)
        self.var_browser = ctk.StringVar(value="Google Chrome / Chromium")
        self.var_direct_dashboard = ctk.BooleanVar(value=False)
        self.var_auto_solve_captcha = ctk.BooleanVar(value=True)
        self.var_tesseract_path = ctk.StringVar(value="")
        self.var_download_dir = ctk.StringVar(value="")
        self.var_timeout_login = ctk.IntVar(value=60)
        self.var_timeout_pdf = ctk.IntVar(value=30)
        self.var_timeout_attachment = ctk.IntVar(value=30)
        self.var_timeout_xls = ctk.IntVar(value=120)

        # Central Timeouts
        self.TIMEOUT_LOGIN = 60000
        self.TIMEOUT_DOWNLOAD_PDF = 30000
        self.TIMEOUT_DOWNLOAD_ATTACHMENT = 30000
        self.TIMEOUT_DOWNLOAD_XLS = 120000
        self.AUTO_OPEN_FOLDER = True

        # State Variables
        self.voucher_excel_file_path = ""
        self.voucher_selected_sheet = ""
        self.paybill_excel_file_path = ""
        self.paybill_selected_sheet = ""
        self.pol_excel_file_path = ""
        self.pol_selected_sheet = ""
        self.current_treasury = sorted(list(TREASURY_IDS.keys()))[0]
        
        self.stop_requested = False
        self.worker_thread = None
        self.downloaded_folders_this_run = []
        self.heartbeat_state = False
        self.post_process_done = False

        # Persistent Playwright/Browser State for Session Reuse
        self.playwright = None
        self.browser = None
        self.browser_context = None
        self.browser_page = None
        
        self._counter_lock = threading.Lock()
        self.success_count = 0
        self.skip_count = 0
        self.fail_count = 0
        
        # Telemetry State variables
        self.browser_used = "UNKNOWN"
        self.run_start_time = 0
        self.captcha_attempts = 0
        self.captcha_success = 0
        self.timeout_count = 0
        self.session_expiry_count = 0

        # Log & Stats buffers for local API
        self.stats_current = "Idle"
        self.stats_progress = 0.0
        self.log_buffer = []
        self.log_buffer_lock = threading.Lock()
        self.api_port = None
        self.api_server = None

        # Set Operator Name directly from bootstrap parameter
        self.operator_name = operator_name

        # Set up Directory Structure
        self.base_path = os.path.join(APP_DIR, 'USER_DATA', self.operator_name)
        self.log_dir = os.path.join(self.base_path, "logs")
        self.download_dir = os.path.join(self.base_path, "downloads")
        os.makedirs(self.log_dir, exist_ok=True)
        os.makedirs(self.download_dir, exist_ok=True)

        self.run_log_file = os.path.join(self.log_dir, f"run_{time.strftime('%Y%m%d_%H%M%S')}.txt")
        self.cred_file = os.path.join(self.base_path, "cred.txt")

        # Load settings
        self.load_config()
        
        # Tab-Specific Output Reports
        self.voucher_master_report = os.path.join(self.download_dir, "Downloaded_Vouchers_Report.txt")
        self.voucher_failed_file = os.path.join(self.download_dir, "Failed_Vouchers.txt")
        self.paybill_master_report = os.path.join(self.download_dir, "PayBill_Downloaded_Report.txt")
        self.paybill_failed_file = os.path.join(self.download_dir, "PayBill_Failed_Vouchers.txt")

        # GUI RENDER
        self._build_gui()
        self.load_credentials()
        self._animate_heartbeat()

        # Log OCR capability status
        self.log("━" * 55)
        self.log("  🌟 IFMIS SUITE AUTOMATION PORTAL LOADED SUCCESSFULLY 🌟")
        self.log(f"  Operator: {self.operator_name} | RUN_ID: {RUN_ID}")
        self.log("━" * 55)
        
        if TESSERACT_AVAILABLE:
            try:
                version = pytesseract.get_tesseract_version()
                self.log(f"✅ Voucher CAPTCHA OCR solver enabled (Tesseract {version})")
            except Exception:
                self.log("⚠️ Tesseract not found - CAPTCHA will require manual entry.")
        else:
            self.log("ℹ️ CAPTCHA will be entered manually.")
            
        start_api_server(self)

    def load_config(self):
        config_path = os.path.join(self.base_path, "config.json")
        defaults = {
            "download_dir": os.path.join(self.base_path, "downloads"),
            "tesseract_path": "",
            "stealth_mode": False,
            "auto_archive": False,
            "dry_run": False,
            "preferred_browser": "Google Chrome / Chromium",
            "direct_dashboard": False,
            "auto_solve_captcha": True,
            "timeout_login": 60,
            "timeout_pdf": 30,
            "timeout_attachment": 45,
            "timeout_xls": 120
        }
        config = defaults.copy()
        if os.path.exists(config_path):
            try:
                with open(config_path, "r", encoding="utf-8") as f:
                    saved = json.load(f)
                    config.update(saved)
            except Exception as e:
                print(f"Error reading config: {e}")
        
        self.download_dir = config["download_dir"]
        self.var_download_dir.set(self.download_dir)
        self.var_tesseract_path.set(config["tesseract_path"])
        self.var_stealth.set(config["stealth_mode"])
        self.var_archive.set(config["auto_archive"])
        self.var_dryrun.set(config["dry_run"])
        self.var_browser.set(config["preferred_browser"])
        self.var_direct_dashboard.set(config["direct_dashboard"])
        self.var_auto_solve_captcha.set(config["auto_solve_captcha"])
        self.var_timeout_login.set(config["timeout_login"])
        self.var_timeout_pdf.set(config["timeout_pdf"])
        self.var_timeout_attachment.set(config["timeout_attachment"])
        self.var_timeout_xls.set(config.get("timeout_xls", 120))
        
        if config["tesseract_path"] and os.path.exists(config["tesseract_path"]):
            if TESSERACT_AVAILABLE:
                pytesseract.pytesseract.tesseract_cmd = config["tesseract_path"]
        
        self.TIMEOUT_LOGIN = config["timeout_login"] * 1000
        self.TIMEOUT_DOWNLOAD_PDF = config["timeout_pdf"] * 1000
        self.TIMEOUT_DOWNLOAD_ATTACHMENT = config["timeout_attachment"] * 1000
        self.TIMEOUT_DOWNLOAD_XLS = config.get("timeout_xls", 120) * 1000

    def save_config(self, show_popup=True):
        config_path = os.path.join(self.base_path, "config.json")
        os.makedirs(os.path.dirname(config_path), exist_ok=True)
        config = {
            "download_dir": self.var_download_dir.get(),
            "tesseract_path": self.var_tesseract_path.get(),
            "stealth_mode": self.var_stealth.get(),
            "auto_archive": self.var_archive.get(),
            "dry_run": self.var_dryrun.get(),
            "preferred_browser": self.var_browser.get(),
            "direct_dashboard": self.var_direct_dashboard.get(),
            "auto_solve_captcha": self.var_auto_solve_captcha.get(),
            "timeout_login": self.var_timeout_login.get(),
            "timeout_pdf": self.var_timeout_pdf.get(),
            "timeout_attachment": self.var_timeout_attachment.get(),
            "timeout_xls": self.var_timeout_xls.get()
        }
        try:
            with open(config_path, "w", encoding="utf-8") as f:
                json.dump(config, f, indent=4)
            self.download_dir = config["download_dir"]
            os.makedirs(self.download_dir, exist_ok=True)
            
            # Update report paths
            self.voucher_master_report = os.path.join(self.download_dir, "Downloaded_Vouchers_Report.txt")
            self.voucher_failed_file = os.path.join(self.download_dir, "Failed_Vouchers.txt")
            self.paybill_master_report = os.path.join(self.download_dir, "PayBill_Downloaded_Report.txt")
            self.paybill_failed_file = os.path.join(self.download_dir, "PayBill_Failed_Vouchers.txt")

            self.TIMEOUT_LOGIN = config["timeout_login"] * 1000
            self.TIMEOUT_DOWNLOAD_PDF = config["timeout_pdf"] * 1000
            self.TIMEOUT_DOWNLOAD_ATTACHMENT = config["timeout_attachment"] * 1000
            self.TIMEOUT_DOWNLOAD_XLS = config["timeout_xls"] * 1000
            
            if config["tesseract_path"] and os.path.exists(config["tesseract_path"]):
                if TESSERACT_AVAILABLE:
                    pytesseract.pytesseract.tesseract_cmd = config["tesseract_path"]
            
            self.log("💾 Application configuration saved successfully.")
            if show_popup:
                messagebox.showinfo("Saved", "Settings Saved Successfully!")
        except Exception as e:
            self.log(f"❌ Failed to save configuration: {e}")
            if show_popup:
                messagebox.showerror("Error", f"Failed to save configuration:\n{e}")

    def save_operator_config(self):
        self.save_config(show_popup=False)

    def browse_download_dir(self):
        path = filedialog.askdirectory(initialdir=self.download_dir)
        if path:
            self.var_download_dir.set(os.path.normpath(path))

    def browse_tesseract_path(self):
        path = filedialog.askopenfilename(filetypes=[("Executable Files", "*.exe")])
        if path:
            self.var_tesseract_path.set(os.path.normpath(path))

    def update_timeout_labels(self, *args):
        try:
            self.lbl_login_timeout.configure(text=f"{self.var_timeout_login.get()}s")
            self.lbl_pdf_timeout.configure(text=f"{self.var_timeout_pdf.get()}s")
            self.lbl_attachment_timeout.configure(text=f"{self.var_timeout_attachment.get()}s")
            self.lbl_xls_timeout.configure(text=f"{self.var_timeout_xls.get()}s")
        except Exception: pass

    def toggle_browser_visibility(self):
        if not self.browser or not self.browser.is_connected():
            self.log("⚠️ No active browser session running to toggle.")
            return
        if sys.platform != "win32":
            self.log("⚠️ Visibility toggle is only supported on Windows.")
            return
        try:
            self.browser_hidden = getattr(self, "browser_hidden", False)
            show_cmd = win32con.SW_SHOW if self.browser_hidden else win32con.SW_HIDE
            action_text = "Showing" if self.browser_hidden else "Hiding"
            
            def window_enum_callback(hwnd, _lParam):
                title = win32gui.GetWindowText(hwnd).lower()
                if any(x in title for x in ["firefox", "chromium", "ifms", "chrome", "edge"]):
                    win32gui.ShowWindow(hwnd, show_cmd)
                return True
                
            win32gui.EnumWindows(window_enum_callback, None)
            self.browser_hidden = not self.browser_hidden
            self.log(f"🕵️ Browser visibility toggle: {action_text} browser window.")
        except Exception as e:
            self.log(f"⚠️ Failed to toggle browser visibility: {e}")

    def _cleanup_browser_session(self):
        page = self.browser_page
        browser = self.browser
        pw = self.playwright
        
        self.browser_page = None
        self.browser_context = None
        self.browser = None
        self.playwright = None
        
        try:
            if page:
                page.close()
        except Exception: pass
        try:
            if browser:
                browser.close()
        except Exception: pass
        try:
            if pw:
                pw.stop()
        except Exception: pass

    def safe_logout_and_close(self, page=None):
        self.log("⏳ Logging out and closing browser session...")
        if page:
            try:
                logout_selectors = [
                    'a[href*="logout"]', 'a[href*="logOut"]', 
                    'input[value*="Logout"]', 'button:has-text("Logout")', 
                    'a:has-text("Logout")', 'a:has-text("logout")',
                    'a:has-text("Exit")', 'a:has-text("exit")'
                ]
                for selector in logout_selectors:
                    if page.locator(selector).is_visible():
                        page.locator(selector).first.click(timeout=3000)
                        self.log("✅ Successfully clicked portal logout button.")
                        page.wait_for_timeout(1000)
                        break
            except Exception:
                pass
        self._cleanup_browser_session()
        self.log("🔒 Browser session closed and cleaned up.")

    def _get_operator_name(self):
        popup = tk.Tk()
        popup.withdraw()
        name = ctk.CTkInputDialog(text="Enter Operator Name:", title="IFMIS Suite Login").get_input()
        popup.destroy()
        if name:
            return "".join(c for c in name if c.isalnum() or c in (" ", "_", "-")).strip()
        return None

    # ==================================================================
    # GUI LAYOUT RENDER
    # ==================================================================
    def _build_gui(self):
        self.configure(fg_color="#0d1117")
        self.title("IFMIS Suite - Developed by AG(A&E) MP Gwalior")
        self.geometry("940x570")
        self.minsize(880, 520)
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._build_left_sidebar()
        self._build_right_panel()

    def _build_left_sidebar(self):
        sidebar = ctk.CTkFrame(self, width=280, corner_radius=0, fg_color="#161b22")
        sidebar.grid(row=0, column=0, sticky="nsew")
        sidebar.grid_rowconfigure(13, weight=1)
        self.sidebar_frame = sidebar

        # Titles
        ctk.CTkLabel(sidebar, text="IFMIS Suite Portal", font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"), text_color="#00c9a7").grid(row=0, column=0, padx=10, pady=(10, 1))
        ctk.CTkLabel(sidebar, text="OFFICE OF THE ACCOUNTANT GENERAL (A&E)", font=ctk.CTkFont(family="Segoe UI", size=9, weight="bold"), text_color="#f8fafc").grid(row=1, column=0, padx=10, pady=(1, 0))
        ctk.CTkLabel(sidebar, text="Madhya Pradesh, Gwalior", font=ctk.CTkFont(family="Segoe UI", size=10), text_color="#8b949e").grid(row=2, column=0, padx=10, pady=(0, 6))

        # Credentials Box
        cred_box = ctk.CTkFrame(sidebar, fg_color="#21262d", corner_radius=12, border_width=1, border_color="#30363d")
        cred_box.grid(row=3, column=0, padx=12, pady=4, sticky="ew")
        cred_box.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(cred_box, text="🔐 IFMS Credentials", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color="#00c9a7").pack(pady=2)
        self.entry_user = ctk.CTkEntry(cred_box, placeholder_text="Username", height=28, fg_color="#0d1117", border_color="#30363d", text_color="#f8fafc", corner_radius=8)
        self.entry_user.pack(padx=10, pady=2, fill="x")
        self.entry_pass = ctk.CTkEntry(cred_box, placeholder_text="Password", show="*", height=28, fg_color="#0d1117", border_color="#30363d", text_color="#f8fafc", corner_radius=8)
        self.entry_pass.pack(padx=10, pady=2, fill="x")
        
        self.btn_save_cred = ctk.CTkButton(
            cred_box, text="Save Credentials", height=26,
            fg_color="transparent", border_width=1, border_color="#00c9a7",
            text_color="#00c9a7", hover_color="#0c362d", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            corner_radius=8,
            command=self.save_credentials,
        )
        self.btn_save_cred.pack(padx=10, pady=(3, 6), fill="x")

        # Global Config Toggles
        toggles_box = ctk.CTkFrame(sidebar, fg_color="transparent")
        toggles_box.grid(row=4, column=0, padx=14, pady=3, sticky="ew")

        self.sw_stealth = ctk.CTkSwitch(toggles_box, text="Hybrid Stealth Mode (Hide Window)", variable=self.var_stealth, font=ctk.CTkFont(family="Segoe UI", size=11), progress_color="#00c9a7", text_color="#f8fafc")
        self.sw_stealth.pack(anchor="w", pady=2)

        self.sw_archive = ctk.CTkSwitch(toggles_box, text="Auto-Archive Downloads to Zip", variable=self.var_archive, font=ctk.CTkFont(family="Segoe UI", size=11), progress_color="#00c9a7", text_color="#f8fafc")
        self.sw_archive.pack(anchor="w", pady=2)

        self.sw_dryrun = ctk.CTkSwitch(toggles_box, text="Dry Run (Bypass IFMS Site/Demo)", variable=self.var_dryrun, font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), progress_color="#d97706", text_color="#d97706")
        self.sw_dryrun.pack(anchor="w", pady=2)

        self.sw_direct_dashboard = ctk.CTkSwitch(toggles_box, text="Direct Dashboard Mode", variable=self.var_direct_dashboard, font=ctk.CTkFont(family="Segoe UI", size=11), progress_color="#00c9a7", text_color="#f8fafc")
        self.sw_direct_dashboard.pack(anchor="w", pady=2)

        self.sw_auto_solve_captcha = ctk.CTkSwitch(toggles_box, text="Auto-Solve CAPTCHA (OCR)", variable=self.var_auto_solve_captcha, font=ctk.CTkFont(family="Segoe UI", size=11), progress_color="#00c9a7", text_color="#f8fafc")
        self.sw_auto_solve_captcha.pack(anchor="w", pady=2)

        # Browser Engine Box
        browser_box = ctk.CTkFrame(sidebar, fg_color="#21262d", corner_radius=12, border_width=1, border_color="#30363d")
        browser_box.grid(row=5, column=0, padx=12, pady=4, sticky="ew")

        ctk.CTkLabel(browser_box, text="Preferred Browser Engine", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color="#00c9a7").pack(pady=2)

        self.opt_browser = ctk.CTkOptionMenu(
            browser_box, variable=self.var_browser,
            values=["Mozilla Firefox (Paybill/POL)", "Google Chrome / Chromium"],
            fg_color="#0d1117", button_color="#21262d", button_hover_color="#30363d",
            text_color="#f8fafc", font=ctk.CTkFont(family="Segoe UI", size=11), dropdown_font=ctk.CTkFont(family="Segoe UI", size=11), height=26, corner_radius=8
        )
        self.opt_browser.pack(padx=10, pady=(1, 6), fill="x")

        # Stats Counters & Unified Progress
        stats_box = ctk.CTkFrame(sidebar, fg_color="#21262d", corner_radius=12, border_color="#30363d", border_width=1)
        stats_box.grid(row=6, column=0, padx=12, pady=4, sticky="ew")

        ctk.CTkLabel(stats_box, text="Task Operations Panel", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color="#8b949e").pack(pady=2)
        
        self.lbl_stats = ctk.CTkLabel(
            stats_box, text="✅ Success: 0   ❌ Failed: 0   ⏭ Skipped: 0",
            font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"), text_color="#f8fafc"
        )
        self.lbl_stats.pack(pady=0)

        self.lbl_current = ctk.CTkLabel(stats_box, text="Current Operation: Idle", font=ctk.CTkFont(family="Segoe UI", size=10), text_color="#8b949e")
        self.lbl_current.pack(pady=0)

        self.progress_bar = ctk.CTkProgressBar(stats_box, height=6, progress_color="#00c9a7", fg_color="#0d1117")
        self.progress_bar.pack(padx=10, pady=6, fill="x")
        self.progress_bar.set(0)

        # Heartbeat animator indicator
        heartbeat_frame = ctk.CTkFrame(stats_box, fg_color="transparent")
        heartbeat_frame.pack(pady=(0, 2))
        ctk.CTkLabel(heartbeat_frame, text="Active Thread Status: ", font=ctk.CTkFont(family="Segoe UI", size=10), text_color="#8b949e").pack(side="left")
        self.lbl_heartbeat = ctk.CTkLabel(heartbeat_frame, text="●", font=ctk.CTkFont(family="Segoe UI", size=12), text_color="gray50")
        self.lbl_heartbeat.pack(side="left")

        # Developer details
        ctk.CTkLabel(
            sidebar, text="Developed for - Indian Audit & Accounts Department",
            font=ctk.CTkFont(family="Segoe UI", size=9), text_color="#8b949e"
        ).grid(row=14, column=0, padx=8, pady=(4, 2))

    def _build_right_panel(self):
        right_panel = ctk.CTkFrame(self, fg_color="transparent")
        right_panel.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
        right_panel.grid_columnconfigure(0, weight=1)
        right_panel.grid_rowconfigure(0, weight=5) # Tabs
        right_panel.grid_rowconfigure(1, weight=3) # Scrolling Console

        # 1. CTkTabview (Top Pane) - Customized to match theme
        self.tabview = ctk.CTkTabview(
            right_panel, corner_radius=12, fg_color="#161b22",
            segmented_button_selected_color="#00c9a7",
            segmented_button_selected_hover_color="#00e5bf",
            segmented_button_unselected_color="#21262d",
            segmented_button_unselected_hover_color="#30363d",
            text_color="#f8fafc",
            command=self.on_tab_switched
        )
        self.tabview.grid(row=0, column=0, sticky="nsew", pady=(0, 6))
        
        self.tabview.add("📄 Vouchers")
        self.tabview.add("💼 Paybills")
        self.tabview.add("📊 POL Reports")
        self.tabview.add("⚙️ Settings")

        self._build_voucher_tab(self.tabview.tab("📄 Vouchers"))
        self._build_paybill_tab(self.tabview.tab("💼 Paybills"))
        self._build_pol_tab(self.tabview.tab("📊 POL Reports"))
        self._build_settings_tab(self.tabview.tab("⚙️ Settings"))

        # 2. Unified Console Output (Bottom Pane)
        console_frame = ctk.CTkFrame(right_panel, fg_color="#161b22", corner_radius=12, border_width=1, border_color="#30363d")
        console_frame.grid(row=1, column=0, sticky="nsew")
        console_frame.grid_columnconfigure(0, weight=1)
        console_frame.grid_columnconfigure(1, weight=0)
        console_frame.grid_columnconfigure(2, weight=0)
        console_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(
            console_frame, text="💻 IFMIS Suite Console Log", 
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), text_color="#00c9a7"
        ).grid(row=0, column=0, pady=4, sticky="w", padx=12)

        # Visibility Toggle & Web Portal Action Buttons (moved here to fit 14-inch screens perfectly)
        self.btn_toggle_browser = ctk.CTkButton(
            console_frame, text="👁 Toggle Browser", height=24, width=120,
            fg_color="transparent", border_width=1, border_color="#00c9a7",
            text_color="#00c9a7", hover_color="#0c362d", font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"),
            corner_radius=8,
            command=self.toggle_browser_visibility
        )
        self.btn_toggle_browser.grid(row=0, column=1, padx=6, pady=4, sticky="e")

        self.btn_launch_portal = ctk.CTkButton(
            console_frame, text="🌐 Launch Web Console", height=24, width=160,
            fg_color="transparent", border_width=1, border_color="#3b82f6",
            text_color="#3b82f6", hover_color="#102a45", font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"),
            corner_radius=8,
            command=self.launch_web_portal
        )
        self.btn_launch_portal.grid(row=0, column=2, padx=(6, 12), pady=4, sticky="e")

        self.output = ctk.CTkTextbox(console_frame, font=ctk.CTkFont(family="Consolas", size=10), fg_color="#090d16", border_width=1, border_color="#30363d", text_color="#f8fafc", corner_radius=8)
        self.output.grid(row=1, column=0, columnspan=3, padx=12, pady=(0, 10), sticky="nsew")
        
        # Configure console tags for colored logs
        self.output.tag_config("success", foreground="#00c9a7")
        self.output.tag_config("skip", foreground="#3b82f6")
        self.output.tag_config("warning", foreground="#fbbf24")
        self.output.tag_config("error", foreground="#f87171")
        self.output.tag_config("info", foreground="#8b949e")

    def on_tab_switched(self):
        selected_tab = self.tabview.get()
        if "Vouchers" in selected_tab:
            self.var_browser.set("Google Chrome / Chromium")
            self.opt_browser.set("Google Chrome / Chromium")
        elif "Paybills" in selected_tab or "POL" in selected_tab:
            self.var_browser.set("Mozilla Firefox (Paybill/POL)")
            self.opt_browser.set("Mozilla Firefox (Paybill/POL)")

    def _build_settings_tab(self, tab):
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(0, weight=1)
        
        scroll_frame = ctk.CTkScrollableFrame(tab, fg_color="transparent")
        scroll_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        scroll_frame.grid_columnconfigure(0, weight=1)
        scroll_frame.grid_columnconfigure(1, weight=3)
        scroll_frame.grid_columnconfigure(2, weight=1)
        
        # Section 1: Paths & Directories
        ctk.CTkLabel(scroll_frame, text="📁 PATHS & DIRECTORIES", font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), text_color="#00c9a7").grid(row=0, column=0, columnspan=3, pady=(5, 10), sticky="w")
        
        # Download directory
        ctk.CTkLabel(scroll_frame, text="Download Root Directory:", font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#f8fafc").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.entry_download_dir = ctk.CTkEntry(scroll_frame, textvariable=self.var_download_dir, height=28, fg_color="#0d1117", border_color="#30363d", text_color="#f8fafc", corner_radius=8)
        self.entry_download_dir.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
        self.btn_browse_download = ctk.CTkButton(
            scroll_frame, text="Browse...", width=80, height=28, 
            fg_color="transparent", border_width=1, border_color="#00c9a7", text_color="#00c9a7", hover_color="#0c362d",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), corner_radius=8,
            command=self.browse_download_dir
        )
        self.btn_browse_download.grid(row=1, column=2, padx=5, pady=5)
        
        # Tesseract directory
        ctk.CTkLabel(scroll_frame, text="Tesseract EXE Path:", font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#f8fafc").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.entry_tesseract_path = ctk.CTkEntry(scroll_frame, textvariable=self.var_tesseract_path, height=28, fg_color="#0d1117", border_color="#30363d", text_color="#f8fafc", corner_radius=8)
        self.entry_tesseract_path.grid(row=2, column=1, sticky="ew", padx=5, pady=5)
        self.btn_browse_tesseract = ctk.CTkButton(
            scroll_frame, text="Browse...", width=80, height=28, 
            fg_color="transparent", border_width=1, border_color="#00c9a7", text_color="#00c9a7", hover_color="#0c362d",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), corner_radius=8,
            command=self.browse_tesseract_path
        )
        self.btn_browse_tesseract.grid(row=2, column=2, padx=5, pady=5)
        
        # Section 2: Timeouts
        ctk.CTkLabel(scroll_frame, text="⏱ TIMEOUT CONFIGURATIONS (SECONDS)", font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), text_color="#00c9a7").grid(row=3, column=0, columnspan=3, pady=(15, 10), sticky="w")
        
        # Login timeout
        ctk.CTkLabel(scroll_frame, text="Portal Login Wait:", font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#f8fafc").grid(row=4, column=0, padx=5, pady=5, sticky="e")
        self.slider_login_timeout = ctk.CTkSlider(scroll_frame, from_=10, to=120, variable=self.var_timeout_login, fg_color="#0d1117", progress_color="#00c9a7", height=15)
        self.slider_login_timeout.grid(row=4, column=1, sticky="ew", padx=5, pady=5)
        self.lbl_login_timeout = ctk.CTkLabel(scroll_frame, text="", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color="#f8fafc")
        self.lbl_login_timeout.grid(row=4, column=2, padx=5, pady=5, sticky="w")
        
        # PDF timeout
        ctk.CTkLabel(scroll_frame, text="PDF Download Timeout:", font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#f8fafc").grid(row=5, column=0, padx=5, pady=5, sticky="e")
        self.slider_pdf_timeout = ctk.CTkSlider(scroll_frame, from_=5, to=60, variable=self.var_timeout_pdf, fg_color="#0d1117", progress_color="#00c9a7", height=15)
        self.slider_pdf_timeout.grid(row=5, column=1, sticky="ew", padx=5, pady=5)
        self.lbl_pdf_timeout = ctk.CTkLabel(scroll_frame, text="", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color="#f8fafc")
        self.lbl_pdf_timeout.grid(row=5, column=2, padx=5, pady=5, sticky="w")
        
        # Attachment timeout
        ctk.CTkLabel(scroll_frame, text="Attachment Timeout:", font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#f8fafc").grid(row=6, column=0, padx=5, pady=5, sticky="e")
        self.slider_attachment_timeout = ctk.CTkSlider(scroll_frame, from_=5, to=120, variable=self.var_timeout_attachment, fg_color="#0d1117", progress_color="#00c9a7", height=15)
        self.slider_attachment_timeout.grid(row=6, column=1, sticky="ew", padx=5, pady=5)
        self.lbl_attachment_timeout = ctk.CTkLabel(scroll_frame, text="", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color="#f8fafc")
        self.lbl_attachment_timeout.grid(row=6, column=2, padx=5, pady=5, sticky="w")
        
        # XLS timeout
        ctk.CTkLabel(scroll_frame, text="XLS Download Timeout:", font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#f8fafc").grid(row=7, column=0, padx=5, pady=5, sticky="e")
        self.slider_xls_timeout = ctk.CTkSlider(scroll_frame, from_=10, to=240, variable=self.var_timeout_xls, fg_color="#0d1117", progress_color="#00c9a7", height=15)
        self.slider_xls_timeout.grid(row=7, column=1, sticky="ew", padx=5, pady=5)
        self.lbl_xls_timeout = ctk.CTkLabel(scroll_frame, text="", font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color="#f8fafc")
        self.lbl_xls_timeout.grid(row=7, column=2, padx=5, pady=5, sticky="w")
        
        # Bind sliders to update labels
        self.var_timeout_login.trace_add("write", self.update_timeout_labels)
        self.var_timeout_pdf.trace_add("write", self.update_timeout_labels)
        self.var_timeout_attachment.trace_add("write", self.update_timeout_labels)
        self.var_timeout_xls.trace_add("write", self.update_timeout_labels)
        self.update_timeout_labels() # Initial labels update
        
        # Action Buttons
        self.btn_save_config = ctk.CTkButton(
            scroll_frame, text="💾 Save Configuration", height=32, 
            fg_color="#00c9a7", hover_color="#00e5bf", text_color="#0d1117", 
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), corner_radius=8,
            command=self.save_config
        )
        self.btn_save_config.grid(row=8, column=0, columnspan=3, pady=(25, 10))

    def _build_voucher_tab(self, tab):
        tab.grid_columnconfigure(0, weight=1)
        
        info = ctk.CTkLabel(
            tab, text="📥 Voucher Downloader (Chromium Engine)\nExtracts Show MPTC PDF, Party Details, and tab attachments via Voucher Excel list.",
            font=ctk.CTkFont(family="Segoe UI", size=11, slant="italic"), text_color="#8b949e"
        )
        info.pack(pady=5)

        # File selection box
        file_box = ctk.CTkFrame(tab, fg_color="#21262d", corner_radius=12, border_width=1, border_color="#30363d")
        file_box.pack(padx=15, pady=8, fill="x")

        self.btn_voucher_excel = ctk.CTkButton(
            file_box, text="📂 Select Voucher Excel List", height=28,
            fg_color="transparent", border_width=1, border_color="#00c9a7", text_color="#00c9a7", hover_color="#0c362d",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), corner_radius=8,
            command=lambda: self.pick_excel_file("voucher")
        )
        self.btn_voucher_excel.pack(pady=(10, 3), padx=15)

        self.lbl_voucher_excel = ctk.CTkLabel(
            file_box, text="No Excel Selected", font=ctk.CTkFont(family="Segoe UI", size=11),
            text_color="#ef4444"
        )
        self.lbl_voucher_excel.pack(pady=(0, 10))

        # Voucher S.No switch
        self.var_voucher_sno = ctk.BooleanVar(value=False)
        self.sw_voucher_sno = ctk.CTkSwitch(
            tab, text="Audit Mode (Auto S.No Injection & Directory Naming)",
            variable=self.var_voucher_sno, font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            progress_color="#00c9a7", text_color="#f8fafc"
        )
        self.sw_voucher_sno.pack(pady=5)

        # Action Buttons
        btn_frame = ctk.CTkFrame(tab, fg_color="transparent")
        btn_frame.pack(pady=10)

        self.btn_voucher_start = ctk.CTkButton(
            btn_frame, text="▶ Start Vouchers", height=32, width=140,
            fg_color="#00c9a7", hover_color="#00e5bf", text_color="#0d1117",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), corner_radius=8,
            command=lambda: self.start_thread("voucher", run_failed=False)
        )
        self.btn_voucher_start.grid(row=0, column=0, padx=6)

        self.btn_voucher_retry = ctk.CTkButton(
            btn_frame, text="🔁 Retry Failed Only", height=32, width=140,
            fg_color="#3b82f6", hover_color="#1d4ed8", text_color="#f8fafc",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), corner_radius=8,
            command=lambda: self.start_thread("voucher", run_failed=True)
        )
        self.btn_voucher_retry.grid(row=0, column=1, padx=6)

        self.btn_voucher_stop = ctk.CTkButton(
            btn_frame, text="🛑 Stop Safely", height=32, width=140,
            fg_color="#ef4444", hover_color="#b91c1c", text_color="#f8fafc",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), corner_radius=8,
            command=self.request_stop
        )
        self.btn_voucher_stop.grid(row=0, column=2, padx=6)

    def _build_paybill_tab(self, tab):
        tab.grid_columnconfigure(0, weight=1)

        info = ctk.CTkLabel(
            tab, text="📥 Paybill Downloader (Firefox Engine)\nRetrieves full employee payroll excel reports using Bill Ref Numbers.",
            font=ctk.CTkFont(family="Segoe UI", size=11, slant="italic"), text_color="#8b949e"
        )
        info.pack(pady=5)

        # File selection box
        file_box = ctk.CTkFrame(tab, fg_color="#21262d", corner_radius=12, border_width=1, border_color="#30363d")
        file_box.pack(padx=15, pady=8, fill="x")

        self.btn_paybill_excel = ctk.CTkButton(
            file_box, text="📂 Select Paybill Excel List", height=28,
            fg_color="transparent", border_width=1, border_color="#00c9a7", text_color="#00c9a7", hover_color="#0c362d",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), corner_radius=8,
            command=lambda: self.pick_excel_file("paybill")
        )
        self.btn_paybill_excel.pack(pady=(10, 3), padx=15)

        self.lbl_paybill_excel = ctk.CTkLabel(
            file_box, text="No Excel Selected", font=ctk.CTkFont(family="Segoe UI", size=11),
            text_color="#ef4444"
        )
        self.lbl_paybill_excel.pack(pady=(0, 10))

        # Paybill S.No switch
        self.var_paybill_sno = ctk.BooleanVar(value=False)
        self.sw_paybill_sno = ctk.CTkSwitch(
            tab, text="Audit Mode (Auto S.No Injection & Directory Naming)",
            variable=self.var_paybill_sno, font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            progress_color="#00c9a7", text_color="#f8fafc"
        )
        self.sw_paybill_sno.pack(pady=5)

        # Action Buttons
        btn_frame = ctk.CTkFrame(tab, fg_color="transparent")
        btn_frame.pack(pady=10)

        self.btn_paybill_start = ctk.CTkButton(
            btn_frame, text="▶ Start Paybills", height=32, width=140,
            fg_color="#00c9a7", hover_color="#00e5bf", text_color="#0d1117",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), corner_radius=8,
            command=lambda: self.start_thread("paybill", run_failed=False)
        )
        self.btn_paybill_start.grid(row=0, column=0, padx=6)

        self.btn_paybill_retry = ctk.CTkButton(
            btn_frame, text="🔁 Retry Failed Only", height=32, width=140,
            fg_color="#3b82f6", hover_color="#1d4ed8", text_color="#f8fafc",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), corner_radius=8,
            command=lambda: self.start_thread("paybill", run_failed=True)
        )
        self.btn_paybill_retry.grid(row=0, column=1, padx=6)

        self.btn_paybill_stop = ctk.CTkButton(
            btn_frame, text="🛑 Stop Safely", height=32, width=140,
            fg_color="#ef4444", hover_color="#b91c1c", text_color="#f8fafc",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), corner_radius=8,
            command=self.request_stop
        )
        self.btn_paybill_stop.grid(row=0, column=2, padx=6)

    def _build_pol_tab(self, tab):
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_columnconfigure(1, weight=1)
        tab.grid_rowconfigure(0, weight=1)

        # LEFT INPUT COLUMN (Date & Parameter fields)
        left_inputs = ctk.CTkFrame(tab, fg_color="transparent")
        left_inputs.grid(row=0, column=0, sticky="nsew", padx=(10, 5), pady=5)
        left_inputs.grid_columnconfigure(1, weight=1)

        # Date Pickers
        ctk.CTkLabel(left_inputs, text="From Date:", font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#f8fafc").grid(row=0, column=0, padx=5, pady=4, sticky="e")
        date_from_frame = ctk.CTkFrame(left_inputs, fg_color="transparent")
        date_from_frame.grid(row=0, column=1, sticky="ew", pady=4)
        date_from_frame.grid_columnconfigure(0, weight=1)
        
        self.entry_pol_from = ctk.CTkEntry(date_from_frame, placeholder_text="DD/MM/YYYY", fg_color="#0d1117", border_color="#30363d", text_color="#f8fafc", height=28, corner_radius=8)
        self.entry_pol_from.grid(row=0, column=0, sticky="ew")
        
        self.btn_pol_from = ctk.CTkButton(
            date_from_frame, text="📅", width=30, height=28, 
            fg_color="transparent", border_width=1, border_color="#00c9a7", text_color="#00c9a7", hover_color="#0c362d",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), corner_radius=8,
            command=lambda: DatePickerModal(self, self.entry_pol_from)
        )
        self.btn_pol_from.grid(row=0, column=1, padx=(5, 0))

        ctk.CTkLabel(left_inputs, text="To Date:", font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#f8fafc").grid(row=1, column=0, padx=5, pady=4, sticky="e")
        date_to_frame = ctk.CTkFrame(left_inputs, fg_color="transparent")
        date_to_frame.grid(row=1, column=1, sticky="ew", pady=4)
        date_to_frame.grid_columnconfigure(0, weight=1)
        
        self.entry_pol_to = ctk.CTkEntry(date_to_frame, placeholder_text="DD/MM/YYYY", fg_color="#0d1117", border_color="#30363d", text_color="#f8fafc", height=28, corner_radius=8)
        self.entry_pol_to.grid(row=0, column=0, sticky="ew")
        
        self.btn_pol_to = ctk.CTkButton(
            date_to_frame, text="📅", width=30, height=28, 
            fg_color="transparent", border_width=1, border_color="#00c9a7", text_color="#00c9a7", hover_color="#0c362d",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), corner_radius=8,
            command=lambda: DatePickerModal(self, self.entry_pol_to)
        )
        self.btn_pol_to.grid(row=0, column=1, padx=(5, 0))

        # Parameters
        ctk.CTkLabel(left_inputs, text="Treasury:", font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#f8fafc").grid(row=2, column=0, padx=5, pady=4, sticky="e")
        self.btn_pol_treasury = ctk.CTkButton(
            left_inputs, text=self.current_treasury, height=28,
            fg_color="#0d1117", border_width=1, border_color="#30363d", text_color="#f8fafc", hover_color="#161b22", anchor="w",
            font=ctk.CTkFont(family="Segoe UI", size=11), corner_radius=8,
            command=self.open_treasury_picker
        )
        self.btn_pol_treasury.grid(row=2, column=1, sticky="ew", pady=4)

        ctk.CTkLabel(left_inputs, text="Bill Type:", font=ctk.CTkFont(family="Segoe UI", size=11), text_color="#f8fafc").grid(row=3, column=0, padx=5, pady=4, sticky="e")
        self.pol_bill_dropdown = ctk.CTkOptionMenu(
            left_inputs, values=["Non-SNA", "SNA"], 
            fg_color="#0d1117", button_color="#21262d", button_hover_color="#30363d", text_color="#f8fafc",
            font=ctk.CTkFont(family="Segoe UI", size=11), dropdown_font=ctk.CTkFont(family="Segoe UI", size=11), height=28, corner_radius=8
        )
        self.pol_bill_dropdown.grid(row=3, column=1, sticky="ew", pady=4)

        # DDO Excel Input (Row 4)
        excel_frame = ctk.CTkFrame(left_inputs, fg_color="#21262d", corner_radius=12, border_width=1, border_color="#30363d")
        excel_frame.grid(row=4, column=0, columnspan=2, pady=5, sticky="ew")
        excel_frame.grid_columnconfigure(0, weight=1)

        self.btn_pol_excel = ctk.CTkButton(
            excel_frame, text="📂 Select DDO Excel List", height=28,
            fg_color="transparent", border_width=1, border_color="#00c9a7", text_color="#00c9a7", hover_color="#0c362d",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), corner_radius=8,
            command=self.pick_pol_excel_file
        )
        self.btn_pol_excel.pack(pady=(10, 3), padx=10)

        self.lbl_pol_excel = ctk.CTkLabel(
            excel_frame, text="No Excel Selected", font=ctk.CTkFont(family="Segoe UI", size=11),
            text_color="#ef4444"
        )
        self.lbl_pol_excel.pack(pady=(0, 10))

        # Action Buttons Frame
        btn_frame = ctk.CTkFrame(left_inputs, fg_color="transparent")
        btn_frame.grid(row=5, column=0, columnspan=2, pady=6, sticky="ew")
        btn_frame.grid_columnconfigure((0, 1), weight=1)

        self.btn_pol_start = ctk.CTkButton(
            btn_frame, text="▶ Start POLs", height=32,
            fg_color="#00c9a7", hover_color="#00e5bf", text_color="#0d1117",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), corner_radius=8,
            command=lambda: self.start_thread("pol", run_failed=False)
        )
        self.btn_pol_start.grid(row=0, column=0, padx=3, pady=2, sticky="ew")

        self.btn_pol_retry = ctk.CTkButton(
            btn_frame, text="🔁 Retry Failed Only", height=32,
            fg_color="#3b82f6", hover_color="#1d4ed8", text_color="#f8fafc",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), corner_radius=8,
            command=lambda: self.start_thread("pol", run_failed=True)
        )
        self.btn_pol_retry.grid(row=0, column=1, padx=3, pady=2, sticky="ew")

        self.btn_pol_stop = ctk.CTkButton(
            btn_frame, text="🛑 Stop Safely", height=32,
            fg_color="#ef4444", hover_color="#b91c1c", text_color="#f8fafc",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), corner_radius=8,
            command=self.request_stop
        )
        self.btn_pol_stop.grid(row=1, column=0, columnspan=2, padx=3, pady=2, sticky="ew")

        # RIGHT INPUT COLUMN (DDO Selection Checkbox Frame)
        right_checklist = ctk.CTkFrame(tab, fg_color="#161b22", border_width=1, border_color="#30363d", corner_radius=12)
        right_checklist.grid(row=0, column=1, sticky="nsew", padx=(5, 10), pady=5)
        right_checklist.grid_columnconfigure(0, weight=1)
        right_checklist.grid_rowconfigure(2, weight=1)

        ctk.CTkLabel(
            right_checklist, text="SELECT DDO CHECKLIST", 
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color="#00c9a7"
        ).grid(row=0, column=0, pady=(6, 2))

        # DDO Search / Filter
        search_box = ctk.CTkFrame(right_checklist, fg_color="#0d1117", border_color="#30363d", border_width=1, height=28, corner_radius=8)
        search_box.grid(row=1, column=0, sticky="ew", padx=8, pady=2)
        search_box.grid_columnconfigure(0, weight=1)
        search_box.grid_propagate(False)

        self.pol_search_entry = ctk.CTkEntry(
            search_box, placeholder_text="🔍 Filter DDOs by name...", 
            border_width=0, fg_color="transparent", text_color="#f8fafc", height=24,
            font=ctk.CTkFont(family="Segoe UI", size=11)
        )
        self.pol_search_entry.grid(row=0, column=0, sticky="ew", padx=4)
        self.pol_search_entry.bind("<KeyRelease>", self.filter_ddos)

        # Scrollable Frame
        self.checklist_frame = ctk.CTkScrollableFrame(right_checklist, fg_color="#0d1117", corner_radius=8)
        self.checklist_frame.grid(row=2, column=0, sticky="nsew", padx=8, pady=(2, 5))

        # Select all box
        self.select_all_frame = ctk.CTkFrame(self.checklist_frame, fg_color="transparent")
        self.select_all_frame.pack(fill="x", padx=8, pady=(3, 3))

        self.select_all_var = ctk.BooleanVar(value=False)
        self.select_all_cb = ctk.CTkCheckBox(
            self.select_all_frame, text="Select All DDOs", variable=self.select_all_var,
            command=self.toggle_all_ddos, fg_color="#00c9a7", hover_color="#00e5bf", text_color="#f8fafc",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold")
        )
        self.select_all_cb.pack(side="left", anchor="w")

        self.lbl_pol_selected_count = ctk.CTkLabel(
            self.select_all_frame, text="(0 selected)", 
            font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"), text_color="#00c9a7"
        )
        self.lbl_pol_selected_count.pack(side="left", padx=8)

        self.ddo_checkboxes = {}
        self.ddo_widgets = {}
        self.refresh_ddo_list(self.current_treasury)

    # ==================================================================
    # TOOL STATE & DECORATIONS
    # ==================================================================
    def _animate_heartbeat(self):
        if self.worker_thread and self.worker_thread.is_alive():
            color = "#F59E0B" if self.var_dryrun.get() else ("#10B981" if self.heartbeat_state else "#047857")
            self.lbl_heartbeat.configure(text_color=color)
            self.heartbeat_state = not self.heartbeat_state
        else:
            self.lbl_heartbeat.configure(text_color="gray50")
        self.after(600, self._animate_heartbeat)

    def set_ui_lock_state(self, is_locked):
        state = "disabled" if is_locked else "normal"
        
        # Sidebar entries
        self.entry_user.configure(state=state)
        self.entry_pass.configure(state=state)
        self.btn_save_cred.configure(state=state)
        self.sw_dryrun.configure(state=state)

        # Voucher Tab
        self.btn_voucher_excel.configure(state=state)
        self.btn_voucher_start.configure(state=state)
        self.btn_voucher_retry.configure(state=state)
        self.sw_voucher_sno.configure(state=state)

        # Paybill Tab
        self.btn_paybill_excel.configure(state=state)
        self.btn_paybill_start.configure(state=state)
        self.btn_paybill_retry.configure(state=state)
        self.sw_paybill_sno.configure(state=state)

        # POL Tab
        self.entry_pol_from.configure(state=state)
        self.entry_pol_to.configure(state=state)
        self.btn_pol_from.configure(state=state)
        self.btn_pol_to.configure(state=state)
        self.btn_pol_treasury.configure(state=state)
        self.pol_bill_dropdown.configure(state=state)
        self.btn_pol_excel.configure(state=state)
        self.btn_pol_start.configure(state=state)
        self.btn_pol_retry.configure(state=state)
        self.select_all_cb.configure(state=state)
        self.pol_search_entry.configure(state=state)
        for cb in self.ddo_widgets.values():
            cb.configure(state=state)

        # Soft colors lock updates
        if is_locked:
            self.btn_voucher_start.configure(fg_color="gray")
            self.btn_voucher_retry.configure(fg_color="gray")
            self.btn_paybill_start.configure(fg_color="gray")
            self.btn_paybill_retry.configure(fg_color="gray")
            self.btn_pol_start.configure(fg_color="gray")
            self.btn_pol_retry.configure(fg_color="gray")
        else:
            self.btn_voucher_start.configure(fg_color="#00c9a7", hover_color="#00e5bf", text_color="#0d1117")
            self.btn_voucher_retry.configure(fg_color="#3b82f6", hover_color="#1d4ed8", text_color="#f8fafc")
            self.btn_paybill_start.configure(fg_color="#00c9a7", hover_color="#00e5bf", text_color="#0d1117")
            self.btn_paybill_retry.configure(fg_color="#3b82f6", hover_color="#1d4ed8", text_color="#f8fafc")
            self.btn_pol_start.configure(fg_color="#00c9a7", hover_color="#00e5bf", text_color="#0d1117")
            self.btn_pol_retry.configure(fg_color="#3b82f6", hover_color="#1d4ed8", text_color="#f8fafc")

    def log(self, message):
        timestamp = time.strftime("%H:%M:%S")
        final_msg = f"[{timestamp}] {message}"
        
        if hasattr(self, "log_buffer_lock"):
            with self.log_buffer_lock:
                self.log_buffer.append(final_msg)
                
        # Determine tag based on content
        tag = None
        lower_msg = message.lower()
        if "✅" in message or "success" in lower_msg:
            tag = "success"
        elif "⏭" in message or "skip" in lower_msg:
            tag = "skip"
        elif "⚠️" in message or "warning" in lower_msg or "verify" in lower_msg:
            tag = "warning"
        elif "❌" in message or "failed" in lower_msg or "error" in lower_msg or "fatal" in lower_msg:
            tag = "error"
        elif "ℹ️" in message or "info" in lower_msg:
            tag = "info"
            
        def append_text(msg=final_msg, tg=tag):
            self.output.insert("end", msg + "\n", tg)
            self.output.see("end")
            
        self.after(0, append_text)
        try:
            with open(self.run_log_file, "a", encoding="utf-8") as f:
                f.write(final_msg + "\n")
        except Exception: pass

    def update_stats(self, current="Idle", progress=None):
        self.stats_current = current
        if progress is not None:
            self.stats_progress = progress
            
        with self._counter_lock:
            s, f, sk = self.success_count, self.fail_count, self.skip_count
        self.after(0, lambda: self.lbl_stats.configure(
            text=f"✅ Success: {s}   ❌ Failed: {f}   ⏭ Skipped: {sk}"
        ))
        self.after(0, lambda: self.lbl_current.configure(text=f"Current: {current} "))
        if progress is not None:
            self.after(0, lambda: self.progress_bar.set(progress))

    def clean_folder_name(self, name):
        for ch in '\\/:*?"<>|=':
            name = str(name).replace(ch, "-")
        return re.sub(r"\s+", "_", name).strip()

    # ==================================================================
    # CREDENTIALS & FILE HANDLING
    # ==================================================================
    def load_credentials(self):
        if not os.path.exists(self.cred_file): return
        try:
            with open(self.cred_file, "r") as f:
                content = f.read().strip()
            if "," in content:
                u, p = content.split(",", 1)
                self.entry_user.insert(0, u)
                self.entry_pass.insert(0, base64.b64decode(p.encode()).decode())
        except Exception: pass

    def save_credentials(self):
        u, p = self.entry_user.get(), self.entry_pass.get()
        try:
            with open(self.cred_file, "w") as f:
                f.write(f"{u},{base64.b64encode(p.encode()).decode()}")
            messagebox.showinfo("Saved", "Credentials Saved Successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save credentials:\n{e}")

    def pick_excel_file(self, mode):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")])
        if not path: return
        try:
            wb = load_workbook(path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            if not sheets: return

            popup = ctk.CTkToplevel(self)
            popup.title("Select Worksheet")
            popup.geometry("380x190")
            popup.configure(fg_color="#0d1117")
            popup.attributes("-topmost", True)
            popup.grab_set()

            ctk.CTkLabel(
                popup, text="Select the worksheet to run:", 
                font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"), text_color="#f8fafc"
            ).pack(pady=(20, 10))
            selected_sheet_var = ctk.StringVar(value=sheets[0])
            
            ctk.CTkOptionMenu(
                popup, variable=selected_sheet_var, values=sheets,
                fg_color="#0d1117", button_color="#21262d", button_hover_color="#30363d", text_color="#f8fafc",
                font=ctk.CTkFont(family="Segoe UI", size=11), dropdown_font=ctk.CTkFont(family="Segoe UI", size=11),
                height=28, corner_radius=8
            ).pack(pady=10)

            def submit_sheet():
                fname = os.path.basename(path)
                display_text = f"Sheet: [{selected_sheet_var.get()}]\nFile: {fname[:22] + '...' if len(fname)>25 else fname}"
                if mode == "voucher":
                    self.voucher_excel_file_path = path
                    self.voucher_selected_sheet = selected_sheet_var.get()
                    self.lbl_voucher_excel.configure(text=display_text, text_color="#00c9a7")
                else:
                    self.paybill_excel_file_path = path
                    self.paybill_selected_sheet = selected_sheet_var.get()
                    self.lbl_paybill_excel.configure(text=display_text, text_color="#00c9a7")
                
                self.log(f"✅ {mode.upper()} Excel Selected: {os.path.basename(path)} [Sheet: {selected_sheet_var.get()}]")
                popup.destroy()

            ctk.CTkButton(
                popup, text="Confirm Selection", command=submit_sheet, 
                fg_color="#00c9a7", hover_color="#00e5bf", text_color="#0d1117",
                font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), corner_radius=8
            ).pack(pady=10)
            self.wait_window(popup)

        except Exception as e:
            self.log(f"❌ Excel Error: {e}")
            messagebox.showerror("Error", f"Failed to read Excel file:\n{e}")

    def pick_pol_excel_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("Excel Files", "*.xls")])
        if not path: return
        try:
            wb = load_workbook(path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            if not sheets: return

            popup = ctk.CTkToplevel(self)
            popup.title("Select Worksheet")
            popup.geometry("380x190")
            popup.configure(fg_color="#0d1117")
            popup.attributes("-topmost", True)
            popup.grab_set()

            ctk.CTkLabel(
                popup, text="Select the worksheet to run:", 
                font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"), text_color="#f8fafc"
            ).pack(pady=(20, 10))
            selected_sheet_var = ctk.StringVar(value=sheets[0])
            
            ctk.CTkOptionMenu(
                popup, variable=selected_sheet_var, values=sheets,
                fg_color="#0d1117", button_color="#21262d", button_hover_color="#30363d", text_color="#f8fafc",
                font=ctk.CTkFont(family="Segoe UI", size=11), dropdown_font=ctk.CTkFont(family="Segoe UI", size=11),
                height=28, corner_radius=8
            ).pack(pady=10)

            def submit_sheet():
                fname = os.path.basename(path)
                display_text = f"Sheet: [{selected_sheet_var.get()}]\nFile: {fname[:22] + '...' if len(fname)>25 else fname}"
                
                self.pol_excel_file_path = path
                self.pol_selected_sheet = selected_sheet_var.get()
                self.lbl_pol_excel.configure(text=display_text, text_color="#00c9a7")
                
                self.log(f"✅ POL Excel Selected: {os.path.basename(path)} [Sheet: {selected_sheet_var.get()}]")
                popup.destroy()
                
                self.auto_check_ddos_from_excel()

            ctk.CTkButton(
                popup, text="Confirm Selection", command=submit_sheet, 
                fg_color="#00c9a7", hover_color="#00e5bf", text_color="#0d1117",
                font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), corner_radius=8
            ).pack(pady=10)
            self.wait_window(popup)

        except Exception as e:
            self.log(f"❌ Excel Error: {e}")
            messagebox.showerror("Error", f"Failed to read Excel file:\n{e}")

    def auto_check_ddos_from_excel(self):
        if not self.pol_excel_file_path: return
        self.log("⏳ Parsing DDO codes from selected Excel sheet...")
        try:
            wb = load_workbook(self.pol_excel_file_path, read_only=True, data_only=True)
            ws = wb[self.pol_selected_sheet]
        except Exception as e:
            self.log(f"❌ Failed to parse Excel: {e}")
            return

        ddo_col_idx = None
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for c_idx, val in enumerate(row):
                if not val or not isinstance(val, str): continue
                text = str(val).lower().strip()
                if any(k in text for k in ['ddo_cd', 'ddo_code', 'ddo code', 'ddocd', 'ddo']):
                    ddo_col_idx = c_idx
                    break
            if ddo_col_idx is not None: break

        if ddo_col_idx is None:
            col_scores = {}
            for col_idx in range(1, min(ws.max_column + 1, 15)):
                score = 0
                for row_idx in range(1, min(ws.max_row + 1, 30)):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val is not None:
                        sval = str(val).strip().split(".")[0]
                        if sval.isdigit() and 4 <= len(sval) <= 8:
                            score += 1
                col_scores[col_idx - 1] = score
            best_col = max(col_scores, key=col_scores.get, default=0)
            if col_scores.get(best_col, 0) > 3:
                ddo_col_idx = best_col

        if ddo_col_idx is None:
            self.log("⚠️ Could not auto-detect DDO column. Checking all columns for potential matches...")
            ddo_col_idx = 0

        ddo_vals = set()
        for row in ws.iter_rows(values_only=True):
            if ddo_col_idx < len(row):
                val = row[ddo_col_idx]
                if val is not None:
                    sval = str(val).strip().split(".")[0]
                    if sval and sval.lower() != "none" and any(c.isdigit() for c in sval):
                        ddo_vals.add(sval)

        wb.close()

        if not ddo_vals:
            self.log("⚠️ No DDO codes found in the Excel sheet.")
            return

        self.log(f"📋 Extracted {len(ddo_vals)} potential DDO code(s) from Excel file.")
        
        matched_count = 0
        for var in self.ddo_checkboxes.values():
            var.set(False)

        for item, var in self.ddo_checkboxes.items():
            item_prefix = item.split("-")[0].strip()
            matched = False
            for val in ddo_vals:
                clean_val = str(val).strip()
                padded_val = clean_val.zfill(len(item_prefix))
                if clean_val == item_prefix or padded_val == item_prefix or clean_val in item or padded_val in item:
                    matched = True
                    break
            if matched:
                var.set(True)
                matched_count += 1

        self.update_selected_count()
        self.log(f"✅ Matched and selected {matched_count} DDO(s) in checklist.")

    # ==================================================================
    # EXCEL COLUMN MAP SEMANTICS & S.NO INJECTION
    # ==================================================================
    def _infer_columns_smart(self, sample_rows):
        best = {
            "mh": (-1, None), "treasury": (-1, None), "voucher": (-1, None),
            "month": (-1, None), "year": (-1, None), "amount": (-1, None),
        }
        max_cols = max(len(r) for r in sample_rows if r)
        for col in range(max_cols):
            values = []
            for row in sample_rows:
                if col < len(row) and row[col] is not None:
                    values.append(str(row[col]).strip())
            if not values: continue
            unique_ratio = len(set(values)) / max(1, len(values))
            
            mh_score = treasury_score = voucher_score = month_score = year_score = amount_score = 0
            for v in values:
                clean = v.upper().replace(".0", "")
                if clean in KNOWN_MAJOR_HEADS: mh_score += 3
                if clean in self.TREASURY_MAP.keys() or clean in self.TREASURY_MAP.values(): treasury_score += 3
                try:
                    iv = int(float(clean))
                    if 1 <= iv <= 12: month_score += 1
                    if 2000 <= iv <= 2100: year_score += 2
                    if iv > 1000: voucher_score += 1
                    if iv > 100: amount_score += 1
                except: pass
            
            voucher_score += unique_ratio * 5
            
            if mh_score > best["mh"][0]: best["mh"] = (mh_score, col)
            if treasury_score > best["treasury"][0]: best["treasury"] = (treasury_score, col)
            if voucher_score > best["voucher"][0]: best["voucher"] = (voucher_score, col)
            if month_score > best["month"][0]: best["month"] = (month_score, col)
            if year_score > best["year"][0]: best["year"] = (year_score, col)
            if amount_score > best["amount"][0]: best["amount"] = (amount_score, col)

        return {k: v[1] for k, v in best.items()}

    def parse_excel_sheet(self, path, sheet_name, audit_mode_enabled):
        self.log(f"⏳ Reading Excel worksheet '{sheet_name}'...")
        file_locked = False
        # Open in read-only mode first because it is 100x faster for large files (14MB+)
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet_name]
        except Exception as e:
            self.log(f"❌ Failed to open workbook: {e}")
            return []

        col_map = {}
        header_row_idx = 1

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
            cmap = {}
            for c_idx, val in enumerate(row):
                if not val or not isinstance(val, str): continue
                text = str(val).lower().strip()
                
                if any(k in text for k in ['s.no', 'sr no', 'sl no', 'serial', 's no']) and 's_no' not in cmap:
                    cmap['s_no'] = c_idx
                elif any(k in text for k in ['mhcd', 'major', 'mh', 'm.h', 'head']) and 'mh' not in cmap:
                    cmap['mh'] = c_idx
                elif text in ['try_cd', 'treasury_cd', 'treasury'] and 'treasury' not in cmap:
                    cmap['treasury'] = c_idx
                elif any(k in text for k in ['treasury', 'try_cd', 'try', 'code']) and 'treasury' not in cmap:
                    if not any(x in text for x in ['try_cdn', 'treasury_cdn', 'trycdn', 'treasurycdn']):
                        cmap['treasury'] = c_idx
                elif any(k in text for k in ['vou_no','sop_vch_no', 'vch', 'voucher', 'vou', 'bill']) and not any(x in text for x in ['dt', 'date', 'typ', 'type', 'cd', 'code', 'amt', 'amount']) and 'voucher' not in cmap:
                    cmap['voucher'] = c_idx
                elif any(k in text for k in ['amt', 'amount', 'net','sop_vch_amt','gross']) and 'amount' not in cmap:
                    cmap['amount'] = c_idx
                elif any(k in text for k in ['month', 'month_in']) and 'month' not in cmap:
                    cmap['month'] = c_idx
                elif any(k in text for k in ['year', 'year_in','yr']) and 'year' not in cmap:
                    cmap['year'] = c_idx
            
            if 'mh' in cmap and 'treasury' in cmap and 'voucher' in cmap:
                col_map = cmap
                header_row_idx = row_idx
                break

        # Fallback profile detection
        if not col_map:
            self.log("⚠️ No headers detected. Applying IFMS raw export detection...")
            try:
                sample = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                treasury_test = str(sample[4]).strip().upper()
                month_test = int(float(sample[5]))
                year_test = int(float(sample[6]))
                mh_test = str(sample[9]).replace(".0", "").strip()
                if treasury_test in self.TREASURY_MAP and 1 <= month_test <= 12 and 2000 <= year_test <= 2100 and mh_test.isdigit() and len(mh_test) == 4:
                    col_map = {'treasury': 4, 'month': 5, 'year': 6, 'mh': 9, 'voucher': 10, 'amount': 21}
                    header_row_idx = 0
                    self.log("✅ Known IFMS raw export structure matched.")
            except Exception: pass

            if not col_map:
                self.log("⚠️ Falling back to smart semantic inference...")
                sample_rows = [row for row in ws.iter_rows(min_row=1, max_row=min(25, ws.max_row), values_only=True)]
                inferred = self._infer_columns_smart(sample_rows)
                
                used_cols = set()
                smart_map = {}
                for key in ['treasury', 'mh', 'voucher', 'month', 'year', 'amount']:
                    val = inferred.get(key)
                    if val is not None and val not in used_cols:
                        smart_map[key] = val
                        used_cols.add(val)
                if 'treasury' in smart_map and 'mh' in smart_map and 'voucher' in smart_map:
                    col_map = smart_map
                    self.log(f"🧠 Inferred column indexes: {col_map}")

        if not col_map:
            self.log("❌ CRITICAL: Failed to identify data columns in Excel sheet.")
            wb.close()
            return []

        # Audit Mode S.No Column Injection
        if audit_mode_enabled and 's_no' not in col_map and not file_locked:
            self.log("⚙️ Injecting 'S.No.' column into Excel file...")
            wb.close() # Close read-only workbook
            
            while True:
                try:
                    wb = load_workbook(path)
                    ws = wb[sheet_name]
                    ws.insert_cols(1)
                    ws.cell(row=header_row_idx, column=1, value="S.No.")
                    s_num = 1
                    new_mh_excel_col = col_map['mh'] + 2
                    for r_idx in range(max(1, header_row_idx + 1), ws.max_row + 1):
                        mh_val = ws.cell(row=r_idx, column=new_mh_excel_col).value
                        if mh_val and str(mh_val).strip() not in ("", "none"):
                            ws.cell(row=r_idx, column=1, value=s_num)
                            s_num += 1
                    wb.save(path)
                    wb.close()
                    self.log("✅ Successfully injected S.No column.")
                    
                    # Reopen in read-only mode for fast data parsing
                    wb = load_workbook(path, read_only=True, data_only=True)
                    ws = wb[sheet_name]
                    
                    for k in list(col_map.keys()): col_map[k] += 1
                    col_map['s_no'] = 0
                    break
                except PermissionError:
                    self.log("⚠️ Permission Denied: Excel file is locked. Prompting user...")
                    ans = messagebox.askretrycancel(
                        "Excel File Locked",
                        f"The spreadsheet '{os.path.basename(path)}' is open in another program (like Microsoft Excel).\n\n"
                        "Please close it in Excel and click 'Retry' to inject serial numbers, or click 'Cancel' to run without saving."
                    )
                    if not ans:
                        self.log("⚠️ S.No injection cancelled. Running in read-only mode.")
                        wb = load_workbook(path, read_only=True, data_only=True)
                        ws = wb[sheet_name]
                        file_locked = True
                        break
                except Exception as e:
                    self.log(f"⚠️ Excel injection error: {e}")
                    wb = load_workbook(path, read_only=True, data_only=True)
                    ws = wb[sheet_name]
                    file_locked = True
                    break

        # Read Data
        data = []
        virtual_sno = 1
        s_no_idx = col_map.get('s_no')
        mh_idx = col_map.get('mh')
        treasury_idx = col_map.get('treasury')
        voucher_idx = col_map.get('voucher')
        amount_idx = col_map.get('amount')
        month_idx = col_map.get('month')
        year_idx = col_map.get('year')

        start_row = header_row_idx + 1 if header_row_idx > 0 else 1
        for row_tuple in ws.iter_rows(min_row=start_row, max_col=ws.max_column, values_only=True):
            def safe_get(idx):
                if idx is not None and idx < len(row_tuple): return row_tuple[idx]
                return None

            mh = safe_get(mh_idx)
            if mh is None or str(mh).strip() in ("", "none"): continue
            
            treasury = safe_get(treasury_idx)
            voucher = safe_get(voucher_idx)
            raw_sno = safe_get(s_no_idx)
            raw_amount = safe_get(amount_idx)
            raw_month = safe_get(month_idx)
            raw_year = safe_get(year_idx)

            if raw_sno is None or str(raw_sno).strip() in ("", "none"): s_no_str = str(virtual_sno)
            else:
                s_no_str = str(raw_sno).strip()
                if s_no_str.endswith(".0"): s_no_str = s_no_str[:-2]
            
            amount_str = str(raw_amount).strip() if raw_amount else ""
            if amount_str.endswith(".0"): amount_str = amount_str[:-2]
            if amount_str.lower() in ("none", ""): amount_str = ""

            voucher_str = str(voucher).strip() if voucher is not None else ""
            if voucher_str.endswith(".0"): voucher_str = voucher_str[:-2]

            folder_name = f"{s_no_str}_{str(mh).strip()}_{str(treasury).strip()}_{voucher_str}"
            if amount_str: folder_name += f"_{amount_str}"

            try: m = int(float(raw_month)) if raw_month else datetime.datetime.now().month
            except: m = datetime.datetime.now().month
            try: y = int(float(raw_year)) if raw_year else datetime.datetime.now().year
            except: y = datetime.datetime.now().year

            if m > 12 and 1 <= y <= 12: m, y = y, m
            if not (1 <= m <= 12): m = datetime.datetime.now().month
            if y < 2000 or y > 2100: y = datetime.datetime.now().year

            data.append({
                "s_no": s_no_str, "mh": str(mh).strip(), "treasury": str(treasury).strip(),
                "voucher": voucher_str, "amount": amount_str, "month": m, "year": y,
                "folder_name": self.clean_folder_name(folder_name),
            })
            virtual_sno += 1

        wb.close()
        return data

    def load_failed_vouchers(self, mode):
        failed_file = self.voucher_failed_file if mode == "voucher" else self.paybill_failed_file
        if not os.path.exists(failed_file): return []
        failed_data = []
        try:
            with open(failed_file, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"): continue
                    parts = line.split("|")
                    if len(parts) == 5: 
                        mh, treasury, voucher, month, year = parts
                        s_no_str, amount_str = "Retry", ""
                    elif len(parts) == 7: 
                        s_no_str, mh, treasury, voucher, amount_str, month, year = parts
                    else: continue
                        
                    folder_name = f"{s_no_str}_{mh}_{treasury}_{voucher}"
                    if amount_str and amount_str != "None": folder_name += f"_{amount_str}"
                    failed_data.append({
                        "s_no": s_no_str, "mh": mh, "treasury": treasury, "voucher": voucher, "amount": amount_str,
                        "month": int(month), "year": int(year), "folder_name": self.clean_folder_name(folder_name),
                    })
        except Exception: pass
        return failed_data

    # ==================================================================
    # POL INPUT HANDLERS
    # ==================================================================
    def open_treasury_picker(self):
        modal = TreasuryPickerModal(self, self.current_treasury, self.set_pol_treasury)
        modal.focus_set()

    def set_pol_treasury(self, treasury_name):
        self.current_treasury = treasury_name
        self.btn_pol_treasury.configure(text=treasury_name)
        self.refresh_ddo_list(treasury_name)

    def refresh_ddo_list(self, treasury_name):
        for widget in self.checklist_frame.winfo_children():
            if widget != self.select_all_frame: widget.destroy()
        
        self.ddo_checkboxes.clear()
        self.ddo_widgets.clear()
        if hasattr(self, 'pol_search_entry'): self.pol_search_entry.delete(0, 'end')
        self.select_all_var.set(False)

        prefix = treasury_name.split("-")[0].strip()
        ddos = ["No DDO Found"]
        for db_key, db_list in DDO_DATABASE.items():
            if db_key.startswith(prefix):
                ddos = db_list
                break

        for ddo in ddos:
            var = ctk.BooleanVar(value=False)
            cb = ctk.CTkCheckBox(
                self.checklist_frame, text=ddo, variable=var,
                command=self.update_selected_count, fg_color="#00c9a7", hover_color="#00e5bf", text_color="#f8fafc",
                font=ctk.CTkFont(family="Segoe UI", size=11)
            )
            cb.pack(anchor="w", padx=10, pady=2)
            self.ddo_checkboxes[ddo] = var
            self.ddo_widgets[ddo] = cb
            
        self.update_selected_count()

    def filter_ddos(self, event=None):
        query = self.pol_search_entry.get().lower()
        for cb in self.ddo_widgets.values(): cb.pack_forget()
        self.select_all_var.set(False)
        
        prefix = self.current_treasury.split("-")[0].strip()
        ddos = ["No DDO Found"]
        for db_key, db_list in DDO_DATABASE.items():
            if db_key.startswith(prefix):
                ddos = db_list
                break
        
        for ddo in ddos:
            if query in ddo.lower() and ddo in self.ddo_widgets:
                self.ddo_widgets[ddo].pack(anchor="w", padx=10, pady=2)

    def toggle_all_ddos(self):
        state = self.select_all_var.get()
        query = self.pol_search_entry.get().lower()
        for ddo, var in self.ddo_checkboxes.items():
            if query in ddo.lower(): var.set(state)
        self.update_selected_count()

    def update_selected_count(self):
        selected_count = sum(1 for var in self.ddo_checkboxes.values() if var.get())
        if hasattr(self, 'lbl_pol_selected_count'):
            self.lbl_pol_selected_count.configure(text=f"({selected_count} selected)")

    # ==================================================================
    # PLAYWRIGHT OCR CAPTCHA SOLVER & DROPDOWNS
    # ==================================================================
    def solve_captcha_ocr(self, page):
        if not TESSERACT_AVAILABLE or not self.var_auto_solve_captcha.get(): 
            return None
        try:
            cmd = getattr(pytesseract.pytesseract, "tesseract_cmd", "tesseract")
            if not cmd:
                cmd = "tesseract"
            import shutil
            resolved_path = cmd if os.path.exists(cmd) else shutil.which(cmd)
            if not resolved_path:
                self.log("⚠️ Tesseract executable path not found or invalid. OCR disabled.")
                return None
            pytesseract.pytesseract.tesseract_cmd = resolved_path

            captcha_img = page.locator("img[src*='captcha'], img[alt*='captcha'], img[alt*='Captcha']").first
            captcha_img.wait_for(state="visible", timeout=5000)
            screenshot_bytes = captcha_img.screenshot()
            
            original_img = Image.open(io.BytesIO(screenshot_bytes))
            
            def run_ocr(proc_img):
                custom_config = r'--psm 7 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789'
                try:
                    text = pytesseract.image_to_string(proc_img, config=custom_config).strip()
                    text = ''.join(c for c in text if c.isalnum())
                    
                    data = pytesseract.image_to_data(proc_img, config=custom_config, output_type=pytesseract.Output.DICT)
                    confidences = [int(conf) for conf in data['conf'] if conf != '-1']
                    avg_conf = sum(confidences) / len(confidences) if confidences else 0
                    return text, avg_conf
                except Exception:
                    return "", 0

            best_text = ""
            best_conf = 0
            
            # Try multiple thresholds to obtain the most confident reading
            thresholds = [120, 140, 165]
            for th in thresholds:
                img = original_img.convert('L')
                width, height = img.size
                img = img.resize((width * 3, height * 3), Image.Resampling.LANCZOS)
                
                enhancer = ImageEnhance.Contrast(img)
                img = enhancer.enhance(3.0)
                img = img.point(lambda p: 255 if p > th else 0)
                img = img.filter(ImageFilter.MedianFilter(size=3))
                
                text, conf = run_ocr(img)
                if len(text) >= 4 and len(text) <= 6:
                    if conf > best_conf:
                        best_text = text
                        best_conf = conf
                        if conf >= 75:
                            break
            
            # Fallback (no thresholding)
            if len(best_text) < 4:
                img = original_img.convert('L')
                width, height = img.size
                img = img.resize((width * 3, height * 3), Image.Resampling.LANCZOS)
                enhancer = ImageEnhance.Contrast(img)
                img = enhancer.enhance(2.0)
                text, conf = run_ocr(img)
                if len(text) >= 4 and len(text) <= 6 and conf > best_conf:
                    best_text = text
                    best_conf = conf

            if best_text and len(best_text) >= 4:
                self.log(f"🔍 OCR CAPTCHA Decoded: '{best_text}' (Confidence: {best_conf:.0f}%)")
                return best_text if best_conf >= 45 else None
            return None
        except Exception as e:
            self.log(f"⚠️ CAPTCHA OCR solver error: {e}")
            return None

    def _perform_portal_login(self, page, context, username, password):
        page.add_init_script("Object.defineProperty(window, 'opener', { get: () => ({}) });")
        self.log("Opening IFMS portal landing page...")
        try:
            page.goto("https://ifmisprod.mptreasury.gov.in/IFMS/", wait_until="domcontentloaded", timeout=60000)
            page.wait_for_load_state("networkidle", timeout=15000)
        except PlaywrightTimeoutError:
            self.timeout_count += 1
            self.log("⚠️ Page load time limits exceeded - attempting takeover...")
        except Exception as e:
            self.log(f"❌ Portal Connection Failed: {e}")
            raise e

        # Detect login page (could be the main page or a popup)
        login_page = None
        for _ in range(20):
            if self.stop_requested: return None
            for context_page in context.pages:
                try:
                    if context_page.locator("input[name='j_username']").count() > 0:
                        login_page = context_page
                        break
                except Exception: pass
            if login_page: break
            time.sleep(1)

        if not login_page:
            self.log("⚠️ Could not detect login fields. Please login manually in the browser window...")
            login_page = page

        # Autofill credentials
        try:
            user_box = login_page.locator("input[name='j_username']")
            user_box.wait_for(state="visible", timeout=10000)
            user_box.click()
            user_box.fill(username)

            pass_box = login_page.locator("input[name='j_password']")
            pass_box.click()
            pass_box.fill(password)
            self.log("✅ Credentials auto-filled.")
        except Exception as e:
            self.log(f"⚠️ Credential filling skipped: {e}")

        # CAPTCHA AUTO-SOLVE
        if TESSERACT_AVAILABLE and self.var_auto_solve_captcha.get():
            self.log("🔐 Attempting automated CAPTCHA solve...")
            self.captcha_attempts += 1
            captcha_text = self.solve_captcha_ocr(login_page)
            if captcha_text:
                try:
                    captcha_input = login_page.locator("input[name='captcha'], input[placeholder*='captcha'], input[placeholder*='Captcha']").first
                    captcha_input.fill(captcha_text)
                    self.captcha_success += 1
                    self.log(f"✅ Filled OCR CAPTCHA: '{captcha_text}'")
                except Exception: pass
        else:
            self.log("👉 Solve the CAPTCHA manually...")

        self.log("⏳ Waiting for user login completion (Solve CAPTCHA + OTP)...")
        
        # Direct Dashboard Mode vs Complex State Tracker
        if self.var_direct_dashboard.get():
            self.log("ℹ️ Direct Dashboard Mode active. Polling for portal landing page directly...")
            while True:
                if self.stop_requested: return None
                
                # Check all pages and frames for dashboard
                for cp in context.pages:
                    try:
                        current_url = cp.url.lower()
                        if "ifms.htm" in current_url and "login.jsp" not in current_url:
                            self.log("✅ Dashboard URL Detected! Taking over...")
                            self._apply_stealth_mode_if_enabled()
                            return cp
                            
                        # Search in page frames
                        for frame in cp.frames:
                            try:
                                if frame.locator("text='Welcome', text='Receipt', text='Disbursement', text='Logout'").count() > 0:
                                    self.log("✅ Dashboard Elements Detected in frame! Taking over...")
                                    self._apply_stealth_mode_if_enabled()
                                    return cp
                            except Exception: pass
                    except Exception: pass
                time.sleep(1)
        else:
            # Full OTP tracking flow
            otp_page_detected = False
            otp_message_shown = False
            login_page_message_shown = False
            last_status = None
            post_otp_wait_start = None
            POST_OTP_TIMEOUT = 30  # seconds to wait after OTP page disappears
            
            while True:
                if self.stop_requested: return None
                
                try:
                    is_login = self._is_on_login_page(login_page)
                    is_otp = self._is_on_otp_page(login_page)
                    is_logged_in = self._is_logged_in(login_page)
                    
                    if is_login:
                        current_status = "login"
                    elif is_otp:
                        current_status = "otp"
                    elif is_logged_in:
                        current_status = "logged_in"
                    else:
                        current_status = "unknown"
                    
                    if current_status != last_status:
                        self.log(f"📍 Page Status: {current_status} | URL: {login_page.url[:65]}...")
                        
                        if current_status == "login" and not login_page_message_shown:
                            self.log("📋 On Login Page - Please enter CAPTCHA and click Login")
                            login_page_message_shown = True
                            post_otp_wait_start = None
                            
                        elif current_status == "otp":
                            if not otp_page_detected:
                                self.log("\n" + "🔔"*20)
                                self.log("📱 OTP PAGE DETECTED!")
                                self.log("👉 Please enter the OTP sent to your registered mobile")
                                self.log("👉 Tool will wait until OTP is verified...")
                                self.log("🔔"*20 + "\n")
                                otp_page_detected = True
                                otp_message_shown = True
                            post_otp_wait_start = None
                                
                        elif current_status == "logged_in":
                            if otp_page_detected:
                                self.log("✅ OTP Verified Successfully!")
                            self.log("✅ Login Complete! Taking over download process...")
                            self._apply_stealth_mode_if_enabled()
                            return login_page
                            
                        elif current_status == "unknown":
                            if otp_page_detected and post_otp_wait_start is None:
                                self.log("⏳ OTP submitted, waiting for page to load...")
                                post_otp_wait_start = time.time()
                                
                        last_status = current_status
                    
                    if is_otp:
                        time.sleep(1)
                        continue
                    
                    if is_logged_in and not is_otp and not is_login:
                        if otp_page_detected:
                            self.log("✅ OTP Verified Successfully!")
                        self.log("✅ Login Complete! Taking over download process...")
                        self._apply_stealth_mode_if_enabled()
                        return login_page
                    
                    if otp_page_detected and current_status == "unknown":
                        if post_otp_wait_start is None:
                            post_otp_wait_start = time.time()
                        elapsed = time.time() - post_otp_wait_start
                        try: login_page.wait_for_load_state("networkidle", timeout=2000)
                        except: pass
                        
                        if self._is_logged_in(login_page):
                            self.log("✅ OTP Verified Successfully!")
                            self.log("✅ Login Complete! Taking over download process...")
                            self._apply_stealth_mode_if_enabled()
                            return login_page
                        
                        if elapsed > POST_OTP_TIMEOUT:
                            self.log(f"⏱️ Timeout after {POST_OTP_TIMEOUT}s - assuming login successful")
                            self.log("✅ Proceeding with download process...")
                            self._apply_stealth_mode_if_enabled()
                            return login_page
                        
                        if int(elapsed) % 5 == 0 and int(elapsed) > 0:
                            remaining = POST_OTP_TIMEOUT - int(elapsed)
                            self.log(f"⏳ Waiting for page... ({remaining}s remaining)")
                    
                except Exception as e:
                    self.log(f"⚠️ Login detection warning: {str(e)[:60]}")
                time.sleep(1)

    def select_dropdown(self, page, selector, code):
        code = str(code).strip().upper()
        if code not in self.TREASURY_MAP: return False
        dropdown = page.locator(f"#{selector}")
        dropdown.wait_for(state="attached")
        try: page.wait_for_function(f"document.querySelectorAll('#{selector} option').length > 1", timeout=5000)
        except PlaywrightTimeoutError: return False

        for opt in dropdown.locator("option").all():
            if opt.inner_text().strip().startswith(self.TREASURY_MAP[code]):
                dropdown.select_option(value=opt.get_attribute("value"))
                return True
        return False

    def _is_on_otp_page(self, page):
        try:
            current_url = page.url.lower()
            otp_url_keywords = ['otp', 'verify', 'verification', 'twofactor', '2fa', 'mfa', 'authenticate']
            if any(keyword in current_url for keyword in otp_url_keywords):
                return True
            
            otp_input_selectors = [
                "input[name*='otp']", "input[name*='OTP']", "input[id*='otp']", "input[id*='OTP']",
                "input[placeholder*='OTP']", "input[placeholder*='otp']", "input[placeholder*='One Time']",
                "input[placeholder*='one time']", "input[placeholder*='Enter OTP']", "input[placeholder*='Enter code']",
                "input[type='tel'][maxlength='6']", "input[type='number'][maxlength='6']", "#txtOTP", "#otp",
                "#otpInput", "#verificationCode",
            ]
            for selector in otp_input_selectors:
                try:
                    if page.locator(selector).count() > 0: return True
                except Exception: pass
            
            try:
                page_text = page.inner_text("body").lower()
                otp_text_indicators = [
                    "enter otp", "enter the otp", "otp sent", "otp has been sent", "verification code",
                    "one time password", "one-time password", "enter verification code", "mobile verification",
                    "sms verification", "otp verification", "verify your mobile", "verify mobile number",
                ]
                if any(indicator in page_text for indicator in otp_text_indicators): return True
            except Exception: pass
            
            otp_button_selectors = [
                "button:has-text('Verify OTP')", "button:has-text('Submit OTP')", "button:has-text('Verify')",
                "input[value*='Verify OTP']", "input[value*='Submit OTP']", "a:has-text('Resend OTP')",
                "button:has-text('Resend OTP')",
            ]
            for selector in otp_button_selectors:
                try:
                    if page.locator(selector).count() > 0: return True
                except Exception: pass
            
            for frame in page.frames:
                try:
                    for selector in otp_input_selectors[:5]:
                        if frame.locator(selector).count() > 0: return True
                except Exception: pass
            return False
        except Exception: return False

    def _is_on_login_page(self, page):
        try:
            current_url = page.url.lower()
            if "login.jsp" in current_url: return True
            if "validatelogin" in current_url: return False
            if page.locator("input[name='j_username']").count() > 0: return True
            if page.locator("input[name='j_password']").count() > 0: return True
            return False
        except Exception: return False

    def _is_logged_in(self, page):
        try:
            current_url = page.url.lower()
            if "login.jsp" in current_url: return False
            if "validatelogin" in current_url: return True
            if "ifms.htm" in current_url: return True
            
            logged_in_selectors = [
                "a[href*='logout']", "text='Logout'", "text='Log Out'", "text='Sign Out'", "#logout", ".logout",
                "text='Welcome'", "text='Dashboard'", "text='Home'", "#menuContainer", "#mainMenu", ".main-menu",
                "nav", "#header", ".header", "text='Receipt'", "text='Disbursement'", "text='Reports'",
                "text='Administration'", "frame[name]",
            ]
            for selector in logged_in_selectors:
                try:
                    if page.locator(selector).count() > 0: return True
                except Exception: pass
            
            # Frame checks for nested portals
            if len(page.frames) > 1:
                return True
            for frame in page.frames:
                try:
                    for selector in logged_in_selectors[:12]:
                        if frame.locator(selector).count() > 0: return True
                except Exception: pass
                
            return False
        except Exception: return False

    def safe_fill(self, page, selector, value):
        loc = page.locator(selector)
        loc.wait_for(state="visible")
        loc.fill(str(value))

    def get_locator_in_frames(self, p_page, selector):
        if p_page.locator(selector).count() > 0: return p_page.locator(selector).first
        for frame in p_page.frames:
            if frame.locator(selector).count() > 0: return frame.locator(selector).first
        return p_page.locator(selector).first

    def logout_session(self, page):
        if not page or page.is_closed(): return False
        self.log("🔓 Logging out from IFMS portal...")
        try:
            for selector in ["a[href*='logout']", "a:has-text('Logout')", "text='Logout'"]:
                try:
                    if page.locator(selector).count() > 0:
                        page.locator(selector).first.click(timeout=3000)
                        page.wait_for_load_state("networkidle", timeout=5000)
                        self.log("✅ Logged out successfully.")
                        return True
                except Exception: continue
            try:
                page.goto("https://ifmisprod.mptreasury.gov.in/IFMS/logout", timeout=5000)
                self.log("✅ Logged out via direct URL.")
                return True
            except Exception: pass
            return False
        except Exception: return False

    # ==================================================================
    # VOUCHER AUTOMATION TASK (Chromium)
    # ==================================================================
    def process_voucher_vch(self, page, context, row_data, failed_rows):
        if self.stop_requested: return
        s_no = row_data["s_no"]
        mh = row_data["mh"]
        treasury = row_data["treasury"]
        vou = row_data["voucher"]
        amount = row_data["amount"]
        month = int(row_data["month"])
        year = int(row_data["year"])
        folder_name = row_data["folder_name"]

        from_date = f"01/{month:02d}/{year}"
        to_date = f"{calendar.monthrange(year, month)[1]:02d}/{month:02d}/{year}"

        folder = os.path.join(self.download_dir, "Voucher", folder_name)
        success_marker = os.path.join(folder, "SUCCESS.marker")

        if os.path.exists(success_marker):
            with self._counter_lock: self.skip_count += 1
            self.update_stats(current=vou)
            return

        os.makedirs(folder, exist_ok=True)
        self.update_stats(current=vou)
        self.log(f"🔍 [{s_no}] Voucher: MH={mh} Try={treasury} Vou={vou}")

        dialog_caught = {"message": None}
        def handle_dialog(dialog):
            dialog_caught["message"] = dialog.message
            dialog.accept()
        page.on("dialog", handle_dialog)

        try:
            page.goto("https://ifmisprod.mptreasury.gov.in/IFMS/ifms.htm?actionFlag=rnd-toApprovedBillsTreasuryWise")
            if "login.jsp" in page.url.lower() or page.locator("input[name='j_username']").count() > 0:
                raise SessionExpiredError("IFMS Session expired - login screen detected.")

            if not self.select_dropdown(page, "cmbTreasuryCodeName", treasury):
                raise Exception("Treasury Code not found in dropdown.")

            self.safe_fill(page, "#txtFrmBillDate", from_date)
            self.safe_fill(page, "#txtToBillDate", to_date)
            self.safe_fill(page, "#txtMajorHead", mh)
            self.safe_fill(page, "#txtSerialNo", vou)
            page.locator("#btnSearch").click()
            page.wait_for_load_state("networkidle")

            if dialog_caught["message"] and "no" in dialog_caught["message"].lower():
                raise Exception("No Results Found")

            bill_link = page.locator("a[onclick*='showBillDtlByCtrlNo']").first
            if not bill_link.is_visible():
                raise Exception("Voucher Details Link not visible.")

            opened_new_tab = False
            try:
                self.log("⏳ Opening voucher details...")
                with context.expect_page(timeout=15000) as new_page_info:
                    bill_link.click()
                bill_page = new_page_info.value
                bill_page.wait_for_load_state("load")
                opened_new_tab = True
            except PlaywrightTimeoutError:
                page.wait_for_load_state("networkidle")
                bill_page = page

            if folder not in self.downloaded_folders_this_run:
                self.downloaded_folders_this_run.append(folder)

            # MPTC PDF
            try:
                mptc_btn = self.get_locator_in_frames(bill_page, "text='Show MPTC'")
                with context.expect_page(timeout=20000) as mptc_page_info: mptc_btn.click()
                mptc_page = mptc_page_info.value
                mptc_page.wait_for_load_state("networkidle")
                client = mptc_page.context.new_cdp_session(mptc_page)
                pdf_data = client.send("Page.printToPDF", {"printBackground": True})
                with open(os.path.join(folder, "Show_MPTC.pdf"), "wb") as f:
                    f.write(base64.b64decode(pdf_data['data']))
                mptc_page.close()
                self.log("✅ MPTC PDF Saved.")
            except Exception as e: self.log(f"⚠️ MPTC skipped: {str(e)[:60]}")

            # Party Details
            try:
                party_btn = self.get_locator_in_frames(bill_page, "text='Download Party Details PDF'")
                with bill_page.expect_download(timeout=self.TIMEOUT_DOWNLOAD_PDF) as download_info: party_btn.click()
                download = download_info.value
                download.save_as(os.path.join(folder, download.suggested_filename))
                self.log("✅ Party Details Saved.")
            except PlaywrightTimeoutError: self.log("⚠️ Party Details PDF timed out.")
            except Exception as e: self.log(f"⚠️ Party Details skipped: {str(e)[:60]}")

            # Attachments Tab
            try:
                tab2_btn = self.get_locator_in_frames(bill_page, "#tab2")
                tab2_btn.click()
                try: bill_page.wait_for_load_state("networkidle", timeout=5000)
                except PlaywrightTimeoutError: pass 

                target_frame = bill_page
                for frame in bill_page.frames:
                    if frame.locator("a[onclick*='openDocumentOnClick']").count() > 0:
                        target_frame = frame
                        break

                links = target_frame.locator("a[onclick*='openDocumentOnClick']").all()
                if links:
                    for idx, link in enumerate(links):
                        try:
                            with bill_page.expect_download(timeout=self.TIMEOUT_DOWNLOAD_ATTACHMENT) as download_info:
                                link.scroll_into_view_if_needed()
                                link.click()
                            download = download_info.value
                            download.save_as(os.path.join(folder, download.suggested_filename))
                            self.log(f"✅ Attachment {idx + 1} Saved.")
                        except Exception as e: self.log(f"⚠️ Attachment {idx + 1} skipped: {str(e)[:60]}")
            except Exception as e: self.log(f"⚠️ Attachments tab error: {str(e)[:60]}")

            if opened_new_tab: bill_page.close()

            with open(success_marker, "w") as f: f.write("DONE")
            with self._counter_lock: self.success_count += 1
            with open(self.voucher_master_report, "a", encoding="utf-8") as report:
                report.write(f"\n[S.NO: {s_no}] Voucher: {vou} | Folder: {folder_name} | Completed: {time.strftime('%H:%M:%S')}\n")

        except SessionExpiredError:
            page.remove_listener("dialog", handle_dialog)
            raise 
        except Exception as e:
            page.remove_listener("dialog", handle_dialog)
            with self._counter_lock: self.fail_count += 1
            failed_rows.append(row_data)
            self.log(f"❌ FAILED: {vou} - {str(e)[:100]}")
            try:
                with open(self.voucher_failed_file, "a", encoding="utf-8") as failed_file:
                    failed_file.write(f"{s_no}|{mh}|{treasury}|{vou}|{amount}|{month}|{year}\n")
            except Exception: pass
        finally:
            page.remove_listener("dialog", handle_dialog)
            try:
                for p in context.pages:
                    if p != page: p.close()
            except Exception: pass


    # ==================================================================
    # PAYBILL AUTOMATION TASK (Firefox)
    # ==================================================================
    def _get_paybill_folder(self, treasury, mh, year, month, s_no, vou):
        treasury_clean = self.clean_folder_name(str(treasury).strip())
        month_label    = f"{year}-{month:02d}"
        voucher_clean  = self.clean_folder_name(str(vou).strip())
        sub_folder     = f"{s_no}_{voucher_clean}"
        return os.path.join(self.download_dir, "PayBill", treasury_clean, month_label, sub_folder)

    def _select_bill_type_paybill(self, page):
        try:
            dd = page.locator("#cmbBillType")
            dd.wait_for(state="visible", timeout=6000)
            for label_try in ["Pay Bill", "PAY BILL", "Pay bill"]:
                try:
                    dd.select_option(label=label_try)
                    self.log(f"✅ Bill Type set to '{label_try}'")
                    return True
                except Exception: pass
            try:
                dd.select_option(value="28~16")
                self.log("✅ Bill Type set to Pay Bill")
                return True
            except Exception: pass
            options = dd.locator("option").all()
            for opt in options:
                txt = opt.inner_text().strip().lower()
                if "pay bill" in txt or "paybill" in txt:
                    dd.select_option(value=opt.get_attribute("value"))
                    self.log(f"✅ Bill Type set to '{opt.inner_text().strip()}'")
                    return True
        except Exception as e:
            self.log(f"⚠️ Bill Type selection failed: {str(e)[:80]}")
        return False

    def _click_bill_ref_no_link(self, page):
        primary = page.locator("a[onclick*='retrievePayrollBillEmpDetails']")
        if primary.count() > 0 and primary.first.is_visible(): return primary.first
        secondary = page.locator("a[onclick*='payrollBillId']")
        if secondary.count() > 0 and secondary.first.is_visible(): return secondary.first
        tertiary = page.locator("a:has-text('Pay Bill/')")
        if tertiary.count() > 0 and tertiary.first.is_visible(): return tertiary.first
        all_links = page.locator("#tblSrchRslt a, table a").all()
        visible_links = [lnk for lnk in all_links if lnk.is_visible()]
        if len(visible_links) >= 2: return visible_links[1]
        return None

    def process_voucher_pb(self, page, context, row_data, failed_rows):
        if self.stop_requested: return
        s_no    = row_data["s_no"]
        mh      = row_data["mh"]
        treasury= row_data["treasury"]
        vou     = row_data["voucher"]
        amount  = row_data["amount"]
        month   = int(row_data["month"])
        year    = int(row_data["year"])

        from_date = f"01/{month:02d}/{year}"
        to_date   = f"{calendar.monthrange(year, month)[1]:02d}/{month:02d}/{year}"

        folder         = self._get_paybill_folder(treasury, mh, year, month, s_no, vou)
        success_marker = os.path.join(folder, "SUCCESS.marker")

        if os.path.exists(success_marker):
            with self._counter_lock: self.skip_count += 1
            self.update_stats(current=vou)
            self.log(f"⏭ [{s_no}] Already downloaded - skipping {vou}")
            return

        os.makedirs(folder, exist_ok=True)
        self.update_stats(current=vou)
        self.log(f"🔍 [{s_no}] Paybill: MH={mh} Try={treasury} Vou={vou}")

        dialog_caught = {"message": None}
        def handle_dialog(dialog):
            dialog_caught["message"] = dialog.message
            dialog.accept()
        page.on("dialog", handle_dialog)

        paybill_page = None
        bill_page = None
        emp_page = None
        try:
            page.goto("https://ifmisprod.mptreasury.gov.in/IFMS/ifms.htm?actionFlag=rnd-toApprovedBillsTreasuryWise", wait_until="domcontentloaded", timeout=30000)
            if "login.jsp" in page.url.lower() or page.locator("input[name='j_username']").count() > 0:
                raise SessionExpiredError("IFMS Session expired - login screen detected.")

            try: page.wait_for_load_state("networkidle", timeout=10000)
            except PlaywrightTimeoutError: pass

            if not self.select_dropdown(page, "cmbTreasuryCodeName", treasury):
                raise Exception(f"Treasury code '{treasury}' not found in dropdown.")

            self.safe_fill(page, "#txtFrmBillDate", from_date)
            self.safe_fill(page, "#txtToBillDate", to_date)
            self.safe_fill(page, "#txtMajorHead", mh)
            self.safe_fill(page, "#txtSerialNo", vou)

            if not self._select_bill_type_paybill(page):
                self.log("⚠️ Bill Type could not be set - proceeding anyway.")

            page.locator("#btnSearch").click()
            try: page.wait_for_load_state("networkidle", timeout=15000)
            except PlaywrightTimeoutError: pass

            if dialog_caught["message"] and "no" in dialog_caught["message"].lower():
                raise Exception("No Records Found")

            page.wait_for_timeout(1000)
            
            # Click the control link to open 1st pop-up
            bill_link = page.locator("a[onclick*='showBillDtlByCtrlNo']").first
            if not bill_link.is_visible():
                bill_link = self._click_bill_ref_no_link(page)

            if bill_link is None or not bill_link.is_visible():
                raise Exception("Bill Control/Ref Link not found or not visible in results table.")

            self.log("⏳ Opening bill details first popup...")
            with context.expect_page(timeout=20000) as new_page_info:
                bill_link.click()
            bill_page = new_page_info.value
            bill_page.wait_for_load_state("load", timeout=30000)
            try: bill_page.wait_for_load_state("networkidle", timeout=10000)
            except PlaywrightTimeoutError: pass
            self.log("✅ Bill Details popup loaded.")

            if folder not in self.downloaded_folders_this_run:
                self.downloaded_folders_this_run.append(folder)

            # Move to 2nd tab inside 1st popup
            self.log("⏳ Navigating to the 2nd tab in the first popup...")
            tab2_btn = self.get_locator_in_frames(bill_page, "#tab2")
            if tab2_btn.count() == 0:
                tab2_btn = bill_page.locator("a:has-text('Paybill'), a:has-text('NPS'), #tab2")
            tab2_btn.first.click()
            try: bill_page.wait_for_load_state("networkidle", timeout=5000)
            except PlaywrightTimeoutError: pass

            # Locate and click 'Employee Wise Bill Details Report' inside the 2nd tab
            self.log("⏳ Locating 'Employee Wise Bill Details Report' button...")
            emp_report_btn = self.get_locator_in_frames(bill_page, "text='Employee Wise Bill Details Report'")
            if emp_report_btn.count() == 0:
                emp_report_btn = self.get_locator_in_frames(bill_page, "a[onclick*='retrievePayrollBillEmpDetails']")
            if emp_report_btn.count() == 0:
                raise Exception("'Employee Wise Bill Details Report' button not found inside 2nd tab of first popup.")

            self.log("⏳ Opening Employee Wise Bill Details Report second popup...")
            with context.expect_page(timeout=25000) as p2_info:
                emp_report_btn.first.click()
            emp_page = p2_info.value
            emp_page.wait_for_load_state("load", timeout=35000)
            try: emp_page.wait_for_load_state("networkidle", timeout=10000)
            except PlaywrightTimeoutError: pass
            self.log("✅ Employee Wise Bill Details Report popup loaded.")

            # Move to 2nd tab of the 2nd popup (Bill Detail)
            self.log("⏳ Navigating to the 2nd tab in the second popup (Bill Detail)...")
            emp_tab2 = self.get_locator_in_frames(emp_page, "a:has-text('Bill Detail')")
            if emp_tab2.count() == 0:
                emp_tab2 = self.get_locator_in_frames(emp_page, "a[rel='tabContent1']")
            if emp_tab2.count() == 0:
                emp_tab2 = self.get_locator_in_frames(emp_page, "#tabmenu li a")
                if emp_tab2.count() >= 2:
                    emp_tab2 = emp_tab2.nth(1)
            if emp_tab2.count() == 0:
                emp_tab2 = emp_page.locator("#tab2, a:has-text('Report')")
            emp_tab2.first.click()
            try: emp_page.wait_for_load_state("networkidle", timeout=5000)
            except PlaywrightTimeoutError: pass

            excel_downloaded = False
            excel_selectors = [
                "a[href*='ExportFormat=4']",
                "a[onclick*='takeAction'][href*='ExportReport']",
                "a:has-text('Excel')",
                "a[href*='excel' i]"
            ]

            excel_btn = None
            for selector in excel_selectors:
                loc = self.get_locator_in_frames(emp_page, selector)
                if loc.count() > 0 and loc.first.is_visible():
                    excel_btn = loc.first
                    break
                    
            if excel_btn is None:
                # Deep search across frames in case of nested frame structure inside second popup
                for frame in emp_page.frames:
                    for selector in excel_selectors:
                        try:
                            floc = frame.locator(selector)
                            if floc.count() > 0 and floc.first.is_visible():
                                excel_btn = floc.first
                                break
                        except Exception: pass
                    if excel_btn: break

            if excel_btn is None:
                raise Exception("Excel export button not found in the employee details report page (second popup, second tab).")

            self.log("📥 Exporting Pay Bill to Excel...")
            with emp_page.expect_download(timeout=self.TIMEOUT_DOWNLOAD_ATTACHMENT) as dl_info:
                excel_btn.click()
            download = dl_info.value

            suggested = download.suggested_filename or "PayBill.xls"
            safe_vou  = self.clean_folder_name(str(vou))
            save_name = f"PayBill_{s_no}_{treasury}_{safe_vou}_{year}{month:02d}{suggested[-4:]}"
            save_path = os.path.join(folder, save_name)
            download.save_as(save_path)
            
            self.log(f"✅ Saved Pay Bill Excel: {save_name} ({os.path.getsize(save_path)/1024:.1f} KB)")
            
            # Inject metadata columns directly into the downloaded Paybill Excel file
            try:
                import pandas as pd
                try:
                    df_list = pd.read_html(save_path)
                    df = df_list[0] if df_list else pd.read_excel(save_path)
                except Exception:
                    df = pd.read_excel(save_path)

                # Insert metadata columns at the beginning of the dataframe
                df.insert(0, "Year", year)
                df.insert(0, "Month", month)
                df.insert(0, "Voucher_No", vou)
                df.insert(0, "Major_Head", mh)
                df.insert(0, "Treasury_Name", treasury)
                df.insert(0, "S_No", s_no)

                df.to_excel(save_path, index=False)
                self.log(f"✅ Injected Excel metadata columns (S_No, Treasury_Name, Major_Head, Voucher_No, Month, Year)")
            except Exception as e:
                self.log(f"⚠️ Could not inject Excel metadata: {e}")

            excel_downloaded = True

            with open(success_marker, "w") as f:
                f.write(f"DONE\nVoucher: {vou}\nSaved: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            with self._counter_lock: self.success_count += 1
            with open(self.paybill_master_report, "a", encoding="utf-8") as report:
                report.write(f"[{time.strftime('%H:%M:%S')}] S.No:{s_no} Voucher:{vou} Try:{treasury} MH:{mh} -> {folder}\n")

        except SessionExpiredError:
            page.remove_listener("dialog", handle_dialog)
            raise
        except Exception as e:
            page.remove_listener("dialog", handle_dialog)
            with self._counter_lock: self.fail_count += 1
            failed_rows.append(row_data)
            self.log(f"❌ FAILED Paybill [{s_no}]: {vou} - {str(e)[:100]}")
            try:
                with open(self.paybill_failed_file, "a", encoding="utf-8") as ff:
                    ff.write(f"{s_no}|{mh}|{treasury}|{vou}|{amount}|{month}|{year}\n")
            except Exception: pass
        finally:
            page.remove_listener("dialog", handle_dialog)
            if bill_page:
                try: bill_page.close()
                except Exception: pass
            if emp_page:
                try: emp_page.close()
                except Exception: pass
            if paybill_page and paybill_page != page:
                try: paybill_page.close()
                except Exception: pass
            try:
                for p in context.pages:
                    if p != page: p.close()
            except Exception: pass


    def _verify_xls_ddo(self, filepath, expected_ddo):
        """
        Verifies if the downloaded XLS file contains the expected DDO code to prevent
        the IFMIS portal caching glitch where reports for another DDO/Treasury are served.
        """
        try:
            if os.path.exists(filepath):
                try:
                    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                    if "<html" in content.lower() or "<table" in content.lower():
                        target = re.sub(r'[^a-zA-Z0-9]', '', str(expected_ddo))
                        normalized_content = re.sub(r'[^a-zA-Z0-9]', '', content)
                        if target in normalized_content:
                            return True
                        else:
                            self.log(f"⚠️ Verification warning: DDO code '{expected_ddo}' not found in HTML XLS file.")
                            return False
                except Exception as text_err:
                    self.log(f"⚠️ Text verification helper error: {text_err}")
                
                if PANDAS_AVAILABLE:
                    try:
                        try: dfs = pd.read_html(filepath)
                        except: dfs = [pd.read_excel(filepath)]
                        
                        target = re.sub(r'[^a-zA-Z0-9]', '', str(expected_ddo))
                        for df in dfs:
                            df_str = re.sub(r'[^a-zA-Z0-9]', '', df.to_string())
                            if target in df_str:
                                return True
                        self.log(f"⚠️ Verification warning: DDO code '{expected_ddo}' not found in parsed Excel sheets.")
                        return False
                    except Exception as pd_err:
                        self.log(f"⚠️ Pandas verification helper error: {pd_err}")
        except Exception as e:
            self.log(f"⚠️ Verification error: {e}")
            
        return True

    # ==================================================================
    # POL AUTOMATION TASK (Firefox)
    # ==================================================================
    def _run_pol_live_mode(self, username, password, target_ddos, run_failed):
        treasury_name = self.current_treasury
        sna_val = "1" if self.pol_bill_dropdown.get() == "Non-SNA" else "0"
        from_date = self.entry_pol_from.get()
        to_date = self.entry_pol_to.get()

        clean_treasury = treasury_name.split("-")[-1].replace(" Treasury", "").replace(" TREASURY", "").replace(" District", "").replace(" Office", "").strip()

        # Split the date range into monthly blocks
        intervals = split_date_range_by_months(from_date, to_date)
        if not intervals:
            self.log("⚠️ Invalid POL date range specified.")
            return

        use_chrome = self.var_browser.get().startswith("Google Chrome")

        # Clean up any existing browser session before starting a fresh run
        self._cleanup_browser_session()

        if use_chrome:
            self.log("🌐 Launching Playwright Engine (Chromium Mode)...")
        else:
            self.log("🌐 Launching Playwright Engine (Firefox Mode)...")

        if not self.playwright:
            self.playwright = sync_playwright().start()

        browser = None
        if use_chrome:
            chrome_args = [
                '--start-maximized',
                '--disable-backgrounding-occluded-windows',
                '--disable-background-timer-throttling',
                '--disable-renderer-backgrounding',
                '--disable-background-networking'
            ]
            try:
                browser = self.playwright.chromium.launch(headless=False, channel="chrome", args=chrome_args)
                self.browser_used = "Google Chrome"
            except Exception:
                try:
                    browser = self.playwright.chromium.launch(headless=False, channel="msedge", args=chrome_args)
                    self.browser_used = "Microsoft Edge"
                except Exception:
                    try:
                        browser = self.playwright.chromium.launch(headless=False, args=chrome_args)
                        self.browser_used = "Bundled Chromium"
                    except Exception as e:
                        self.log(f"❌ Failed to launch Chromium/Chrome: {e}")
                        return
        else:
            try:
                firefox_prefs = {
                    "widget.windows.window_occlusion_tracking.enabled": False,
                    "dom.min_background_timeout_value": 10000,
                    "privacy.reduceTimerPrecision": False,
                }
                browser = self.playwright.firefox.launch(headless=False, firefox_user_prefs=firefox_prefs, args=['-disable-security-sandbox', '-disable-content-sandbox'])
                self.browser_used = "Mozilla Firefox"
            except Exception as e:
                self.log(f"❌ Failed to launch Firefox: {e}")
                return

        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        self.browser = browser
        self.browser_context = context
        self.browser_page = page

        try:
            logged_in_page = self._perform_portal_login(page, context, username, password)
            if not logged_in_page:
                self._cleanup_browser_session()
                return
            page = logged_in_page
            self.browser_page = page
            
            report_url = "https://ifmisprod.mptreasury.gov.in/IFMS/ifms.htm?actionFlag=rnd-paymentOrderListForAG&elementId=10072358"
            
            # Loop over monthly intervals
            for sub_idx, (sub_from, sub_to) in enumerate(intervals):
                if self.stop_requested: break
                
                try:
                    d, m, y = sub_from.split('/')
                    month_year = f"{m}_{y}"
                except ValueError:
                    month_year = sub_from.replace('/', '_')

                # POL target directory inside operator downloads
                download_dir = os.path.join(self.download_dir, f"POL_{clean_treasury}_{month_year}")
                os.makedirs(download_dir, exist_ok=True)
                if download_dir not in self.downloaded_folders_this_run:
                    self.downloaded_folders_this_run.append(download_dir)

                success_path = os.path.join(download_dir, f"Success_Summary_{clean_treasury}_{month_year}.txt")
                failed_path = os.path.join(download_dir, f"Failed_Summary_{clean_treasury}_{month_year}.txt")

                # Compute target DDOs for this specific month
                if not run_failed:
                    month_ddos = target_ddos
                    with open(success_path, "w", encoding="utf-8") as f:
                        f.write(f"=== SUCCESS & SKIPPED SUMMARY ({treasury_name}) ===\n")
                    with open(failed_path, "w", encoding="utf-8") as f: pass
                else:
                    month_ddos = []
                    if os.path.exists(failed_path):
                        with open(failed_path, "r", encoding="utf-8") as f:
                            for line in f:
                                ddo = line.strip()
                                if ddo: month_ddos.append(ddo)
                    if not month_ddos:
                        self.log(f"⏭️ No failed POL DDOs for month {month_year} to retry.")
                        continue
                    
                    with open(success_path, "a", encoding="utf-8") as f:
                        f.write(f"\n=== RETRY RUN ({time.strftime('%H:%M:%S')}) ===\n")
                    with open(failed_path, "a", encoding="utf-8") as f: pass

                self.log(f"\n📅 Processing Date Block: {sub_from} to {sub_to} ({sub_idx+1}/{len(intervals)})")
                
                # Navigate to report page
                try:
                    page.goto(report_url, timeout=30000)
                except Exception as e:
                    self.log(f"⚠️ Failed navigating to POL report URL: {e}. Trying reload...")
                    page.goto(report_url)

                total_ddos = len(month_ddos)
                current_failed_ddos = []

                for idx, ddo_name in enumerate(month_ddos):
                    if self.stop_requested: break
                    clean_ddo = re.sub(r"[^\w\s-]", "", ddo_name.split("-")[-1])[:20].strip()
                    expected_filename = os.path.join(download_dir, f"{clean_treasury}_{clean_ddo}_{month_year}.xls")
                    
                    if os.path.exists(expected_filename):
                        self.log(f"⏭️ Skipping: {clean_ddo} (Already Downloaded)")
                        with open(success_path, "a", encoding="utf-8") as f:
                            f.write(f"⏭️ SKIPPED: {ddo_name} (Already Downloaded)\n")
                        with self._counter_lock: self.skip_count += 1
                        
                        progress_val = (idx + 1) / total_ddos
                        self.update_stats(current=clean_ddo, progress=progress_val)
                        continue

                    self.log(f"🚀 POL: Processing [{idx+1}/{total_ddos}] - {ddo_name}")
                    
                    ddo_code_prefix = ddo_name.split("-")[0].strip()
                    max_attempts = 3
                    download_success = False
                    
                    for attempt in range(1, max_attempts + 1):
                        if self.stop_requested: break
                        if attempt > 1:
                            self.log(f"🔄 Retrying DDO {clean_ddo} due to portal/verification glitch (Attempt {attempt}/{max_attempts})...")
                            try: page.locator("button:has-text('Back')").click()
                            except:
                                try: page.locator("#back").click()
                                except: page.goto(report_url)
                            page.wait_for_timeout(1000)

                        try:
                            page.wait_for_selector("#tsryList", timeout=15000)
                        except:
                            if page.locator("input[name='j_username']").count() > 0:
                                raise SessionExpiredError("IFMS Session expired - login screen detected.")
                            self.log("🔄 Portal glitch - reloading report window...")
                            page.goto(report_url)
                            try: page.wait_for_selector("#tsryList", timeout=30000)
                            except: raise Exception("Failed to load treasury dropdown.")

                        page.locator("#tsryList").select_option(value=TREASURY_IDS[treasury_name])
                        
                        ddo_selected = False
                        for _ in range(50): 
                            if self.stop_requested: break
                            try:
                                is_selected = page.evaluate(f"""
                                    () => {{
                                        let select = document.querySelector('#ddoCode');
                                        if (!select || select.options.length === 0) return false;
                                        for (let i = 0; i < select.options.length; i++) {{
                                            let opt = select.options[i];
                                            if (opt.value.includes('{ddo_code_prefix}') || opt.text.includes('{ddo_code_prefix}')) {{
                                                select.value = opt.value; 
                                                select.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                                return true;
                                            }}
                                        }}
                                        return false;
                                    }}
                                """)
                                if is_selected:
                                    ddo_selected = True
                                    break 
                            except: pass 
                            page.wait_for_timeout(200)

                        if self.stop_requested: break

                        if not ddo_selected:
                            self.log(f"⚠️ DDO prefix not found: {ddo_name}")
                            break

                        try:
                            page.locator("#fromDate").fill(sub_from, force=True)
                            page.locator("#toDate").fill(sub_to, force=True)
                        except:
                            page.evaluate(f"""
                                document.getElementById('fromDate').value = '{sub_from}';
                                document.getElementById('toDate').value = '{sub_to}';
                            """)

                        page.locator("#snaFlag").select_option(value=sna_val)
                        page.locator("#btnGenReport").evaluate("node => node.click()")
                        
                        try: page.wait_for_selector("#statusbar", state="hidden", timeout=30000)
                        except: pass
                        page.wait_for_timeout(500)

                        no_records = False
                        if page.locator("text=/No records found./i").count() > 0:
                            no_records = True
                        else:
                            try:
                                page.wait_for_selector("button:has-text('Download XLS')", timeout=4000)
                            except Exception:
                                no_records = True

                        if no_records:
                            self.log(f"⏭️ No records or report found for DDO '{clean_ddo}' - skipping.")
                            with open(success_path, "a", encoding="utf-8") as f:
                                f.write(f"⏭️ SKIPPED: {ddo_name} (No Records/Report Found.)\n")
                            with self._counter_lock: self.skip_count += 1
                            download_success = True
                            
                            try: page.locator("button:has-text('Back')").click()
                            except:
                                try: page.locator("#back").click()
                                except: page.goto(report_url)
                            break

                        try:
                            page.wait_for_selector("#po-page-size", timeout=5000)
                            page.locator("#po-page-size").select_option(value="9999")
                            self.log("🔄 Selected 'Groups per page' to 'All'.")
                            page.wait_for_timeout(1000)
                        except:
                            self.log("⚠️ Custom table sizing dropdown not found - exporting default view.")

                        try:
                            self.log("⏳ Exporting report XLS...")
                            with page.expect_download(timeout=self.TIMEOUT_DOWNLOAD_XLS) as download_info:
                                page.locator("button:has-text('Download XLS')").click()
                            download = download_info.value
                            download.save_as(expected_filename)
                            
                            if self._verify_xls_ddo(expected_filename, ddo_code_prefix):
                                self.log(f"✅ Saved verified XLS file: {os.path.basename(expected_filename)}")
                                with open(success_path, "a", encoding="utf-8") as f: f.write(f"✅ SUCCESS: {ddo_name}\n")
                                with self._counter_lock: self.success_count += 1
                                download_success = True
                                
                                try: page.locator("button:has-text('Back')").click()
                                except:
                                    try: page.locator("#back").click()
                                    except: page.goto(report_url)
                                break
                            else:
                                self.log(f"❌ Verification failed for DDO '{ddo_code_prefix}'. Report data belongs to another Treasury/DDO (IFMIS portal glitch). Retrying...")
                                if os.path.exists(expected_filename):
                                    try: os.remove(expected_filename)
                                    except Exception: pass
                        except Exception as e:
                            self.log(f"❌ Failed to download XLS: {str(e).splitlines()[0][:50]}")

                    if self.stop_requested: break

                    if not download_success:
                        self.log(f"❌ DDO download failed after {max_attempts} attempts: {ddo_name}")
                        current_failed_ddos.append(ddo_name)
                        with open(failed_path, "a", encoding="utf-8") as f: f.write(f"{ddo_name}\n")
                        with self._counter_lock: self.fail_count += 1
                        
                        try: page.locator("button:has-text('Back')").click()
                        except:
                            try: page.locator("#back").click()
                            except: page.goto(report_url)

                    progress_val = (idx + 1) / total_ddos
                    self.update_stats(current=clean_ddo, progress=progress_val)
                    page.wait_for_timeout(500)

                # Universal master excel merge (using pandas)
                xls_files = [os.path.join(download_dir, f) for f in os.listdir(download_dir) if f.endswith('.xls')]
                if xls_files and PANDAS_AVAILABLE:
                    self.log(f"⏳ Merging all exported POL sheets for {month_year} into Master Excel database...")
                    try:
                        dfs = []
                        for fp in xls_files:
                            try: df = pd.read_html(fp)[0]
                            except:
                                try: df = pd.read_excel(fp)
                                except: continue
                            df['Source_DDO_File'] = os.path.basename(fp)
                            dfs.append(df)
                        if dfs:
                            master_df = pd.concat(dfs, ignore_index=True)
                            master_filename = f"Master_file_{clean_treasury}_{month_year}.xlsx"
                            master_df.to_excel(os.path.join(download_dir, master_filename), index=False)
                            self.log(f"🌟 Combined Master sheet built: {master_filename}")
                    except Exception as e:
                        self.log(f"⚠️ Master merge failed: {e}")
                elif xls_files:
                    self.log("ℹ️ Master excel merge skipped - Pandas library is not installed.")

                self.log(f"🎯 POL Batch Completed for Date Block: {sub_from} to {sub_to}")
                if current_failed_ddos:
                    self.log(f"⚠️ DDO failures logged ({len(current_failed_ddos)}) for {month_year}.")
                else:
                    self.log(f"✨ POL downloads for {month_year} completed with 100% success rate!")

        finally:
            self.safe_logout_and_close(page)
            self.log("👉 The UI is unlocked! You can run failed ones or start a new batch directly.")

    # ==================================================================
    # CORE THREAD RUNNER SEQUENCE
    # ==================================================================
    def start_thread(self, mode, run_failed=False):
        if self.worker_thread and self.worker_thread.is_alive(): return

        # Inputs verification
        username = self.entry_user.get().strip()
        password = self.entry_pass.get().strip()
        
        if not username or not password:
            messagebox.showerror("Missing Input", "Please enter your IFMS Username and Password in the sidebar.")
            return

        if mode == "voucher" and not run_failed and not self.voucher_excel_file_path:
            messagebox.showerror("Missing Input", "Please select a Voucher Excel spreadsheet before starting.")
            return
            
        if mode == "paybill" and not run_failed and not self.paybill_excel_file_path:
            messagebox.showerror("Missing Input", "Please select a Paybill Excel spreadsheet before starting.")
            return

        if mode == "pol" and not run_failed:
            if not self.entry_pol_from.get() or not self.entry_pol_to.get():
                messagebox.showerror("Missing Input", "Please specify both 'From' and 'To' dates for POL.")
                return
            selected_ddos = [name for name, var in self.ddo_checkboxes.items() if var.get()]
            if not selected_ddos:
                messagebox.showerror("Missing Input", "Please select at least one DDO from the checklist.")
                return

        self.stop_requested = False
        self.post_process_done = False
        self.run_start_time = time.time()
        
        # Telemetry cleanups
        self.captcha_attempts = 0
        self.captcha_success = 0
        self.timeout_count = 0
        self.session_expiry_count = 0

        with self._counter_lock: self.success_count = self.skip_count = self.fail_count = 0
        self.downloaded_folders_this_run = []

        self.update_stats(progress=0.0)
        self.set_ui_lock_state(is_locked=True)

        self.worker_thread = threading.Thread(target=self._run_automation_task, args=(mode, run_failed), daemon=True)
        self.worker_thread.start()

    def _run_automation_task(self, mode, run_failed):
        # Prevent Windows system from going to sleep during active run
        if sys.platform == "win32":
            try:
                import ctypes
                ctypes.windll.kernel32.SetThreadExecutionState(0x80000000 | 0x00000001 | 0x00000002)
            except Exception: pass

        asyncio.set_event_loop(asyncio.new_event_loop())
        try:
            username = self.entry_user.get().strip()
            password = self.entry_pass.get().strip()

            if mode == "pol":
                # POL download workflow
                target_ddos = [name for name, var in self.ddo_checkboxes.items() if var.get()]
                if not target_ddos and not run_failed:
                    self.log("❌ No DDOs selected.")
                    return
                
                if self.var_dryrun.get():
                    self._run_dry_mode_pol(target_ddos, run_failed)
                else:
                    self._run_pol_live_mode(username, password, target_ddos, run_failed)
            else:
                # Excel-based workflows (Voucher / Paybill)
                if run_failed:
                    excel_rows = self.load_failed_vouchers(mode)
                else:
                    path = self.voucher_excel_file_path if mode == "voucher" else self.paybill_excel_file_path
                    sheet = self.voucher_selected_sheet if mode == "voucher" else self.paybill_selected_sheet
                    audit_mode = self.var_voucher_sno.get() if mode == "voucher" else self.var_paybill_sno.get()
                    excel_rows = self.parse_excel_sheet(path, sheet, audit_mode)

                if not excel_rows:
                    self.log("❌ No valid rows found in selected worksheet database.")
                    return

                total_items = len(excel_rows)
                failed_file = self.voucher_failed_file if mode == "voucher" else self.paybill_failed_file
                if not run_failed:
                    try:
                        with open(failed_file, "w", encoding="utf-8") as f:
                            f.write(f"# Failed {mode.upper()} Vouchers - Started {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                            f.write("# Format: S.No|MH|Treasury|Voucher|Amount|Month|Year\n")
                    except Exception: pass

                if self.var_dryrun.get():
                    self._run_dry_mode_excel(excel_rows, total_items, mode)
                else:
                    self._run_excel_live_mode(username, password, excel_rows, total_items, mode)

            if not self.post_process_done:
                self._post_process(mode)

        except Exception as e:
            if self.stop_requested:
                self.log("🛑 Run aborted: Stop safety triggered by user.")
            else:
                self.log(f"❌ FATAL COMPILER EXCEPTION: {e}")
                try:
                    with open(self.run_log_file, "a", encoding="utf-8") as f:
                        f.write(traceback.format_exc())
                except Exception: pass
        finally:
            if sys.platform == "win32":
                try:
                    import ctypes
                    # Keep persistent sleep prevention active for the main GUI app window,
                    # so we re-assert the global execution state instead of resetting to 0x80000000.
                    ctypes.windll.kernel32.SetThreadExecutionState(0x80000000 | 0x00000001 | 0x00000002)
                except Exception: pass
            self.after(0, lambda: self.set_ui_lock_state(is_locked=False))

    # ==================================================================
    # EXCEL LIVE PORTAL SEQUENCER (Vouchers & Paybills)
    # ==================================================================
    def _run_excel_live_mode(self, username, password, excel_rows, total_items, mode):
        use_chrome = (mode == "voucher") or self.var_browser.get().startswith("Google Chrome")
        
        # Clean up any existing browser session before starting a fresh run
        self._cleanup_browser_session()

        self.log(f"🌐 Launching Playwright browser interface ({'Chromium' if use_chrome else 'Firefox'} Mode)...")
        if not self.playwright:
            self.playwright = sync_playwright().start()

        browser = None
        if use_chrome:
            chrome_args = [
                '--start-maximized',
                '--disable-backgrounding-occluded-windows',
                '--disable-background-timer-throttling',
                '--disable-renderer-backgrounding',
                '--disable-background-networking'
            ]
            try:
                browser = self.playwright.chromium.launch(headless=False, channel="chrome", args=chrome_args)
                self.browser_used = "Google Chrome"
            except Exception:
                try:
                    browser = self.playwright.chromium.launch(headless=False, channel="msedge", args=chrome_args)
                    self.browser_used = "Microsoft Edge"
                except Exception:
                    try:
                        browser = self.playwright.chromium.launch(headless=False, args=chrome_args)
                        self.browser_used = "Bundled Chromium"
                    except Exception: pass
            if not browser:
                self.log("❌ Chromium/Chrome/Edge browser engines are not found. Launching standard Firefox fallback...")
                try:
                    firefox_prefs = {
                        "widget.windows.window_occlusion_tracking.enabled": False,
                        "dom.min_background_timeout_value": 10000,
                        "privacy.reduceTimerPrecision": False,
                    }
                    browser = self.playwright.firefox.launch(headless=False, firefox_user_prefs=firefox_prefs, args=['-disable-security-sandbox', '-disable-content-sandbox'])
                    self.browser_used = "Mozilla Firefox Fallback"
                except Exception:
                    self.log("❌ Failed to launch any browser. Please install Playwright browsers or Chrome.")
                    return
        else:
            # Paybill is natively Firefox Mode
            try:
                firefox_prefs = {
                    "widget.windows.window_occlusion_tracking.enabled": False,
                    "dom.min_background_timeout_value": 10000,
                    "privacy.reduceTimerPrecision": False,
                }
                browser = self.playwright.firefox.launch(headless=False, firefox_user_prefs=firefox_prefs, args=['-disable-security-sandbox', '-disable-content-sandbox'])
                self.browser_used = "Mozilla Firefox"
            except Exception as e:
                self.log(f"❌ Failed to launch Firefox for Paybill: {e}")
                return

        context_args = {"accept_downloads": True}
        if use_chrome:
            context_args["no_viewport"] = True
        context = browser.new_context(**context_args)
        page = context.new_page()

        self.browser = browser
        self.browser_context = context
        self.browser_page = page

        try:
            if True:
                logged_in_page = self._perform_portal_login(page, context, username, password)
                if not logged_in_page:
                    self._cleanup_browser_session()
                    return
                page = logged_in_page
                self.browser_page = page

                # Automation Execution loop
                failed_rows = []
                session_expired = False
                
                for idx, row_data in enumerate(excel_rows, 1):
                    if self.stop_requested: break
                    self.log(f"📊 Progress: {idx}/{total_items}")
                    self.update_stats(current=row_data['voucher'], progress=(idx / total_items))
                    
                    try:
                        if mode == "voucher":
                            self.process_voucher_vch(page, context, row_data, failed_rows)
                        else:
                            self.process_voucher_pb(page, context, row_data, failed_rows)
                    except SessionExpiredError:
                        self.session_expiry_count += 1
                        self.log("🔴 IFMS portal session expired. Process sequence terminated.")
                        session_expired = True
                        break

                # Retrying fail loops
                if failed_rows and not self.stop_requested and not session_expired:
                    self.log("\n🔁 SEQUENCE INITIATING SECOND RUN OF FAILURES...")
                    retry_rows, failed_rows = failed_rows.copy(), []
                    for idx, row_data in enumerate(retry_rows, 1):
                        if self.stop_requested: break
                        self.update_stats(current=f"Retry: {row_data['voucher']}", progress=(idx / len(retry_rows)))
                        try:
                            if mode == "voucher":
                                self.process_voucher_vch(page, context, row_data, failed_rows)
                            else:
                                self.process_voucher_pb(page, context, row_data, failed_rows)
                        except SessionExpiredError:
                            session_expired = True
                            break

                self.log("🎯 Batch execution sequence finished.")
                self.update_stats(current="Idle", progress=1.0)
                
                with self._counter_lock: total_failed = self.fail_count
                if total_failed > 0: self.log(f"📝 failures logged under: Failed_{mode.upper()}s.txt")
                
                self._post_process(mode)
                self.post_process_done = True 

        finally:
            self.safe_logout_and_close(page)
            self.log("👉 The UI is unlocked! You can run failed ones or start a new batch directly.")

    # ==================================================================
    # DRY SIMULATOR WORKFLOWS (Offline Tests)
    # ==================================================================
    def _run_dry_mode_excel(self, excel_rows, total_items, mode):
        self.log("\n" + "="*55)
        self.log(f"🚀 DRY SEQUENCE INITIATED: BYPASSING LIVE IFMS PORTAL ({mode.upper()})")
        self.log("="*55)
        for idx, row_data in enumerate(excel_rows, 1):
            if self.stop_requested: break
            vou = row_data["voucher"]
            if mode == "voucher":
                folder = os.path.join(self.download_dir, "Voucher", row_data["folder_name"])
            else:
                folder = self._get_paybill_folder(row_data["treasury"], row_data["mh"], row_data["year"], row_data["month"], row_data["s_no"], vou)

            self.log(f"📊 Simulated Process: [{idx}/{total_items}] Paybill: {vou}")
            self.update_stats(current=f"Mocking {vou}", progress=(idx / total_items))
            time.sleep(1.2)

            os.makedirs(folder, exist_ok=True)
            if folder not in self.downloaded_folders_this_run:
                self.downloaded_folders_this_run.append(folder)

            with open(os.path.join(folder, "SUCCESS.marker"), "w") as f:
                f.write(f"DONE\nVoucher: {vou}\nDry Run: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            with self._counter_lock: self.success_count += 1

        self.log("\n🎯 SIMULATED BATCH COMPLETE")
        self.update_stats(current="Idle", progress=1.0)
        with self._counter_lock: total_simulated = self.success_count
        self.log(f"✅ Simulated {total_simulated} {mode.upper()} download trees successfully.")
        self._post_process(mode)
        self.post_process_done = True

    def _run_dry_mode_pol(self, target_ddos, run_failed):
        self.log("\n" + "="*55)
        self.log("🚀 DRY SEQUENCE INITIATED: BYPASSING LIVE IFMS PORTAL (POL)")
        self.log("="*55)
        
        treasury_name = self.current_treasury
        from_date = self.entry_pol_from.get()
        to_date = self.entry_pol_to.get()
        
        clean_treasury = treasury_name.split("-")[-1].replace(" Treasury", "").replace(" TREASURY", "").replace(" District", "").replace(" Office", "").strip()

        intervals = split_date_range_by_months(from_date, to_date)
        if not intervals:
            self.log("⚠️ Invalid POL date range specified.")
            return

        for sub_idx, (sub_from, sub_to) in enumerate(intervals):
            if self.stop_requested: break
            
            try:
                d, m, y = sub_from.split('/')
                month_year = f"{m}_{y}"
            except ValueError:
                month_year = sub_from.replace('/', '_')

            download_dir = os.path.join(self.download_dir, f"POL_{clean_treasury}_{month_year}")
            os.makedirs(download_dir, exist_ok=True)
            if download_dir not in self.downloaded_folders_this_run:
                self.downloaded_folders_this_run.append(download_dir)

            failed_path = os.path.join(download_dir, f"Failed_Summary_{clean_treasury}_{month_year}.txt")
            
            if not run_failed:
                month_ddos = target_ddos
                with open(failed_path, "w", encoding="utf-8") as f: pass
            else:
                month_ddos = []
                if os.path.exists(failed_path):
                    with open(failed_path, "r", encoding="utf-8") as f:
                        for line in f:
                            ddo = line.strip()
                            if ddo: month_ddos.append(ddo)
                if not month_ddos:
                    self.log(f"⏭️ No simulated failed POL DDOs for {month_year} to retry.")
                    continue
                with open(failed_path, "w", encoding="utf-8") as f: pass

            self.log(f"\n📅 Simulated Date Block: {sub_from} to {sub_to} ({sub_idx+1}/{len(intervals)})")
            total_ddos = len(month_ddos)
            
            for idx, ddo_name in enumerate(month_ddos):
                if self.stop_requested: break
                clean_ddo = re.sub(r"[^\w\s-]", "", ddo_name.split("-")[-1])[:20].strip()
                self.log(f"📊 Simulated Process: [{idx+1}/{total_ddos}] DDO: {clean_ddo}")
                self.update_stats(current=clean_ddo, progress=((idx+1)/total_ddos))
                time.sleep(0.5)

                expected_filename = os.path.join(download_dir, f"{clean_treasury}_{clean_ddo}_{month_year}.xls")
                with open(expected_filename, "w", encoding="utf-8") as f:
                    f.write("<table><tr><td>MOCK DATA TABLE</td></tr></table>")
                with self._counter_lock: self.success_count += 1

            # Master merge simulation
            xls_files = [os.path.join(download_dir, f) for f in os.listdir(download_dir) if f.endswith('.xls')]
            if xls_files and PANDAS_AVAILABLE:
                self.log(f"⏳ Simulated merge of sheets into Master Excel database for {month_year}...")
                try:
                    dfs = []
                    for fp in xls_files:
                        df = pd.read_html(fp)[0]
                        df['Source_DDO_File'] = os.path.basename(fp)
                        dfs.append(df)
                    if dfs:
                        master_df = pd.concat(dfs, ignore_index=True)
                        master_filename = f"Master_file_{clean_treasury}_{month_year}.xlsx"
                        master_df.to_excel(os.path.join(download_dir, master_filename), index=False)
                except: pass

        self.log("\n🎯 SIMULATED BATCH COMPLETE")
        self.update_stats(current="Idle", progress=1.0)
        with self._counter_lock: total_simulated = self.success_count
        self.log(f"✅ Simulated {total_simulated} POL sheet downloads successfully.")
        self._post_process("pol")
        self.post_process_done = True

    # ==================================================================
    # POST-PROCESSING, HIDING & EXITS
    # ==================================================================
    def _apply_stealth_mode_if_enabled(self):
        if not self.var_stealth.get() or sys.platform != "win32": return
        self.log("🕵️ Relocating Browser process execution into the background...")
        try:
            def window_enum_callback(hwnd, _lParam):
                title = win32gui.GetWindowText(hwnd).lower()
                if any(x in title for x in ["firefox", "chromium", "ifms", "chrome", "edge"]): 
                    win32gui.ShowWindow(hwnd, win32con.SW_HIDE)
                return True
            win32gui.EnumWindows(window_enum_callback, None)
        except Exception: pass

    def _post_process(self, mode):
        # 1. Zip Archiving
        if self.var_archive.get() and self.downloaded_folders_this_run:
            self.log("📦 Compressing completed downloads into Zip archive...")
            zip_filename = os.path.join(self.download_dir, f"IFMIS_{mode.upper()}_Batch_{time.strftime('%Y%m%d_%H%M%S')}.zip")
            try:
                with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for folder in self.downloaded_folders_this_run:
                        if not os.path.exists(folder): continue
                        if os.path.isfile(folder):
                            zipf.write(folder, os.path.basename(folder))
                        else:
                            for root, _, files in os.walk(folder):
                                for file in files:
                                    file_path = os.path.join(root, file)
                                    zipf.write(file_path, os.path.relpath(file_path, self.download_dir))
                self.log(f"✅ Zip Archive Saved: {os.path.basename(zip_filename)}")
            except Exception as e: self.log(f"⚠️ Archive compression failed: {e}")

        # 2. Run completion Telemetry Logging
        runtime_payload = {
            "run_id": RUN_ID,
            "event": "batch_complete",
            "mode": mode.upper(),
            "operator": self.operator_name,
            "success": self.success_count,
            "failed": self.fail_count,
            "skipped": self.skip_count,
            "runtime_seconds": int(time.time() - self.run_start_time),
            "browser_used": self.browser_used,
            "captcha_attempts": self.captcha_attempts,
            "captcha_success": self.captcha_success,
            "timeouts": self.timeout_count,
            "session_expiry_count": self.session_expiry_count,
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        threading.Thread(
            target=send_silent_log,
            args=(self.operator_name, json.dumps(runtime_payload)),
            daemon=True
        ).start()

        # 3. Auto open folders
        if self.AUTO_OPEN_FOLDER:
            try:
                target_folder = self.download_dir
                if mode == "pol" and self.downloaded_folders_this_run:
                    target_folder = self.downloaded_folders_this_run[0]
                elif mode == "paybill":
                    target_folder = os.path.join(self.download_dir, "PayBill")
                    os.makedirs(target_folder, exist_ok=True)
                
                if sys.platform == "win32": os.startfile(target_folder)
                elif sys.platform == "darwin": subprocess.Popen(["open", target_folder])
                else: subprocess.Popen(["xdg-open", target_folder])
            except Exception: pass

    def request_stop(self):
        self.stop_requested = True
        self.log("🛑 STOP Safety Button Triggered. Terminating active sessions...")
        
        is_browser_open = False
        try:
            if self.browser and self.browser.is_connected():
                is_browser_open = True
        except Exception:
            pass

        if is_browser_open:
            self.log("⏳ Browser is open. Attempting to log out and close browser context...")
            def perform_stop():
                try:
                    if self.browser_page:
                        self.safe_logout_and_close(self.browser_page)
                except Exception:
                    pass
                finally:
                    self._cleanup_browser_session()
                    self.stop_requested = False
                    self.after(0, lambda: self.set_ui_lock_state(is_locked=False))

            threading.Thread(target=perform_stop, daemon=True).start()
        else:
            self.log("ℹ️ Browser is already closed. Resetting program to reuse...")
            self._cleanup_browser_session()
            self.stop_requested = False
            self.set_ui_lock_state(is_locked=False)

    def launch_web_portal(self):
        if getattr(self, "api_port", None):
            import webbrowser
            webbrowser.open(f"http://127.0.0.1:{self.api_port}/")
            self.log(f"🌐 Opened Web Portal in your default browser at http://127.0.0.1:{self.api_port}/")
        else:
            self.log("❌ Integrated API Server is not running.")

    def run_scraper_via_api(self, mode, username, password, config_params):
        def run_on_main():
            self.entry_user.delete(0, 'end')
            self.entry_user.insert(0, username)
            self.entry_pass.delete(0, 'end')
            self.entry_pass.insert(0, password)
            
            if "stealth_mode" in config_params:
                self.var_stealth.set(config_params["stealth_mode"])
            if "direct_dashboard" in config_params:
                self.var_direct_dashboard.set(config_params["direct_dashboard"])
            if "auto_solve_captcha" in config_params:
                self.var_auto_solve_captcha.set(config_params["auto_solve_captcha"])
            if "timeout_login" in config_params:
                self.var_timeout_login.set(config_params["timeout_login"])
                self.TIMEOUT_LOGIN = config_params["timeout_login"] * 1000
            if "timeout_pdf" in config_params:
                self.var_timeout_pdf.set(config_params["timeout_pdf"])
                self.TIMEOUT_DOWNLOAD_PDF = config_params["timeout_pdf"] * 1000
            if "timeout_attachment" in config_params:
                self.var_timeout_attachment.set(config_params["timeout_attachment"])
                self.TIMEOUT_DOWNLOAD_ATTACHMENT = config_params["timeout_attachment"] * 1000
            if "timeout_xls" in config_params:
                self.var_timeout_xls.set(config_params["timeout_xls"])
                self.TIMEOUT_DOWNLOAD_XLS = config_params["timeout_xls"] * 1000

            if mode == "voucher":
                if "voucher_excel_file_path" in config_params:
                    self.voucher_excel_file_path = config_params["voucher_excel_file_path"]
            elif mode == "paybill":
                if "paybill_excel_file_path" in config_params:
                    self.paybill_excel_file_path = config_params["paybill_excel_file_path"]
            elif mode == "pol":
                if "pol_from" in config_params:
                    self.entry_pol_from.delete(0, 'end')
                    self.entry_pol_from.insert(0, config_params["pol_from"])
                if "pol_to" in config_params:
                    self.entry_pol_to.delete(0, 'end')
                    self.entry_pol_to.insert(0, config_params["pol_to"])
                
                if "selected_ddos" in config_params:
                    target_ddos = config_params["selected_ddos"]
                    for ddo, var in self.ddo_checkboxes.items():
                        var.set(False)
                    for ddo in target_ddos:
                        if ddo in self.ddo_checkboxes:
                            self.ddo_checkboxes[ddo].set(True)
                        else:
                            for k in self.ddo_checkboxes.keys():
                                if ddo in k:
                                    self.ddo_checkboxes[k].set(True)
                                    break
                    self._update_pol_selection_label()

            self.start_thread(mode)
            
        self.after(0, run_on_main)

    def on_closing(self):
        self.request_stop()
        try:
            if self.browser:
                self.browser.close()
        except: pass
        try:
            if self.playwright:
                self.playwright.stop()
        except: pass
        if sys.platform == "win32":
            try:
                import ctypes
                ctypes.windll.kernel32.SetThreadExecutionState(0x80000000)
            except Exception: pass
        self.destroy()


# =========================================================
# WEB PORTAL HTTP API SERVER
# =========================================================
import http.server
import json
import urllib.parse
from http.server import ThreadingHTTPServer

if getattr(sys, "frozen", False):
    WEB_DIST_DIR = os.path.join(sys._MEIPASS, "web_dist")
else:
    WEB_DIST_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "web_dist")

class PortalAPIHandler(http.server.BaseHTTPRequestHandler):
    app = None
    
    def log_message(self, format, *args):
        pass

    def end_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(204)
        self.end_headers()

    def do_GET(self):
        parsed_url = urllib.parse.urlparse(self.path)
        path = parsed_url.path
        
        if path == "/api/download/status":
            self.send_json({
                "is_running": self.app.worker_thread is not None and self.app.worker_thread.is_alive(),
                "success_count": getattr(self.app, "success_count", 0),
                "skip_count": getattr(self.app, "skip_count", 0),
                "fail_count": getattr(self.app, "fail_count", 0),
                "stop_requested": getattr(self.app, "stop_requested", False),
                "current_action": getattr(self.app, "stats_current", "Idle"),
                "progress": getattr(self.app, "stats_progress", 0.0),
                "run_time": int(time.time() - self.app.run_start_time) if getattr(self.app, "run_start_time", 0) > 0 and self.app.worker_thread and self.app.worker_thread.is_alive() else 0
            })
            return
            
        elif path == "/api/download/logs":
            with self.app.log_buffer_lock:
                logs = list(self.app.log_buffer)
                self.app.log_buffer.clear()
            self.send_json({"logs": logs})
            return

        elif path == "/api/config":
            self.send_json({
                "username": self.app.entry_user.get().strip(),
                "stealth_mode": self.app.var_stealth.get(),
                "direct_dashboard": self.app.var_direct_dashboard.get(),
                "auto_solve_captcha": self.app.var_auto_solve_captcha.get(),
                "timeout_login": self.app.var_timeout_login.get(),
                "timeout_pdf": self.app.var_timeout_pdf.get(),
                "timeout_attachment": self.app.var_timeout_attachment.get(),
                "timeout_xls": self.app.var_timeout_xls.get(),
                "download_dir": self.app.var_download_dir.get(),
                "tesseract_path": self.app.var_tesseract_path.get(),
                "known_ddos": list(self.app.ddo_checkboxes.keys())
            })
            return

        elif path == "/api/files/list":
            download_root = self.app.var_download_dir.get()
            if not download_root:
                download_root = os.path.join(APP_DIR, "USER_DATA", self.app.entry_user.get().strip() or "nitin")
            
            files_list = []
            if os.path.exists(download_root):
                for root_dir, dirs, filenames in os.walk(download_root):
                    for filename in filenames:
                        if filename.endswith((".xlsx", ".xls", ".csv")):
                            full_path = os.path.join(root_dir, filename)
                            rel_path = os.path.relpath(full_path, download_root)
                            category = "unknown"
                            if "voucher" in root_dir.lower():
                                category = "voucher"
                            elif "paybill" in root_dir.lower():
                                category = "paybill"
                            elif "pol" in root_dir.lower():
                                category = "pol"
                                
                            stat = os.stat(full_path)
                            files_list.append({
                                "name": filename,
                                "rel_path": rel_path,
                                "full_path": full_path,
                                "category": category,
                                "size": stat.st_size,
                                "mtime": stat.st_mtime
                            })
            self.send_json({"files": files_list})
            return

        elif path == "/api/files/content":
            query = urllib.parse.parse_qs(parsed_url.query)
            file_path = query.get("path", [None])[0]
            if not file_path:
                self.send_error(400, "Missing path parameter")
                return
                
            if not os.path.exists(file_path):
                self.send_error(404, "File not found")
                return
                
            if not file_path.endswith((".xlsx", ".xls", ".csv")):
                self.send_error(403, "Access denied")
                return
                
            try:
                self.send_response(200)
                self.send_header("Content-Type", "application/octet-stream")
                self.send_header("Content-Disposition", f"attachment; filename=\"{os.path.basename(file_path)}\"")
                stat = os.stat(file_path)
                self.send_header("Content-Length", str(stat.st_size))
                self.end_headers()
                
                with open(file_path, "rb") as f:
                    self.wfile.write(f.read())
            except Exception as e:
                self.send_error(500, f"Error reading file: {e}")
            return

        # Serve static files
        clean_path = path.lstrip('/')
        if clean_path == "":
            clean_path = "index.html"
            
        file_to_serve = os.path.join(WEB_DIST_DIR, clean_path)
        if not os.path.exists(file_to_serve) or os.path.isdir(file_to_serve):
            file_to_serve = os.path.join(WEB_DIST_DIR, "index.html")
            
        if not os.path.exists(file_to_serve):
            self.send_response(404)
            self.send_header("Content-Type", "text/plain")
            self.end_headers()
            self.wfile.write(b"Static assets not found.")
            return
            
        content_type = "text/html"
        if file_to_serve.endswith(".js"):
            content_type = "application/javascript"
        elif file_to_serve.endswith(".css"):
            content_type = "text/css"
        elif file_to_serve.endswith(".png"):
            content_type = "image/png"
        elif file_to_serve.endswith(".ico"):
            content_type = "image/x-icon"
        elif file_to_serve.endswith(".svg"):
            content_type = "image/svg+xml"
            
        try:
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            stat = os.stat(file_to_serve)
            self.send_header("Content-Length", str(stat.st_size))
            self.end_headers()
            
            with open(file_to_serve, "rb") as f:
                self.wfile.write(f.read())
        except Exception as e:
            self.send_error(500, f"Server error: {e}")

    def do_POST(self):
        parsed_url = urllib.parse.urlparse(self.path)
        path = parsed_url.path
        
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length) if content_length > 0 else b""
        
        if path == "/api/download/start":
            try:
                params = json.loads(post_data.decode('utf-8'))
                mode = params.get("mode")
                username = params.get("username")
                password = params.get("password")
                config_params = params.get("config", {})
                
                if not mode or not username or not password:
                    self.send_json({"success": False, "error": "Missing mode, username, or password"}, 400)
                    return
                
                if self.app.worker_thread and self.app.worker_thread.is_alive():
                    self.send_json({"success": False, "error": "Download run is already in progress"}, 400)
                    return
                
                self.app.run_scraper_via_api(mode, username, password, config_params)
                self.send_json({"success": True, "message": f"{mode.upper()} download thread scheduled"})
            except Exception as e:
                self.send_json({"success": False, "error": str(e)}, 500)
            return
            
        elif path == "/api/download/stop":
            self.app.after(0, self.app.request_stop)
            self.send_json({"success": True, "message": "Stop request flagged and session cleanup triggered"})
            return

        elif path == "/api/config":
            try:
                params = json.loads(post_data.decode('utf-8'))
                def apply_config():
                    if "download_dir" in params:
                        self.app.var_download_dir.set(params["download_dir"])
                    if "tesseract_path" in params:
                        self.app.var_tesseract_path.set(params["tesseract_path"])
                    if "stealth_mode" in params:
                        self.app.var_stealth.set(params["stealth_mode"])
                    if "direct_dashboard" in params:
                        self.app.var_direct_dashboard.set(params["direct_dashboard"])
                    if "auto_solve_captcha" in params:
                        self.app.var_auto_solve_captcha.set(params["auto_solve_captcha"])
                    if "timeout_login" in params:
                        self.app.var_timeout_login.set(params["timeout_login"])
                        self.app.TIMEOUT_LOGIN = params["timeout_login"] * 1000
                    if "timeout_pdf" in params:
                        self.app.var_timeout_pdf.set(params["timeout_pdf"])
                        self.app.TIMEOUT_DOWNLOAD_PDF = params["timeout_pdf"] * 1000
                    if "timeout_attachment" in params:
                        self.app.var_timeout_attachment.set(params["timeout_attachment"])
                        self.app.TIMEOUT_DOWNLOAD_ATTACHMENT = params["timeout_attachment"] * 1000
                    if "timeout_xls" in params:
                        self.app.var_timeout_xls.set(params["timeout_xls"])
                        self.app.TIMEOUT_DOWNLOAD_XLS = params["timeout_xls"] * 1000
                    self.app.save_operator_config()
                self.app.after(0, apply_config)
                self.send_json({"success": True, "message": "Configuration updated"})
            except Exception as e:
                self.send_json({"success": False, "error": str(e)}, 500)
            return

        elif path == "/api/files/upload":
            try:
                query = urllib.parse.parse_qs(parsed_url.query)
                filename = query.get("name", ["uploaded_file.xlsx"])[0]
                mode = query.get("mode", ["voucher"])[0]
                
                upload_dir = os.path.join(self.app.base_path, "uploads")
                os.makedirs(upload_dir, exist_ok=True)
                target_path = os.path.join(upload_dir, filename)
                
                with open(target_path, "wb") as f:
                    f.write(post_data)
                
                def set_upload_path():
                    if mode == "voucher":
                        self.app.voucher_excel_file_path = target_path
                    elif mode == "paybill":
                        self.app.paybill_excel_file_path = target_path
                self.app.after(0, set_upload_path)
                
                self.send_json({"success": True, "path": target_path})
            except Exception as e:
                self.send_json({"success": False, "error": str(e)}, 500)
            return

        self.send_error(404, "Not Found")

    def send_json(self, data, status=200):
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        response_bytes = json.dumps(data).encode('utf-8')
        self.send_header("Content-Length", str(len(response_bytes)))
        self.end_headers()
        self.wfile.write(response_bytes)

def start_api_server(app_instance):
    port = 5000
    server = None
    while port < 5050:
        try:
            PortalAPIHandler.app = app_instance
            server = ThreadingHTTPServer(('127.0.0.1', port), PortalAPIHandler)
            break
        except OSError:
            port += 1
            
    if server:
        app_instance.api_port = port
        app_instance.log(f"🌐 Integrated Web Server started on http://127.0.0.1:{port}")
        t = threading.Thread(target=server.serve_forever, daemon=True)
        t.start()
        app_instance.api_server = server
    else:
        app_instance.log("❌ Failed to start integrated API server (no ports available)")


# =========================================================
# SYSTEM EXECUTIONS BOOTSTRAP
# =========================================================
if __name__ == "__main__":
    # high-DPI scaling bypasses
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except: pass

    # Initialize lightweight splash screen immediately using standard Tkinter
    import tkinter as tk
    from tkinter import ttk
    
    splash = tk.Tk()
    splash.title("IFMIS Suite Startup")
    splash.geometry("400x230")
    splash.configure(bg="#0d1117")
    
    # Center splash window
    screen_width = splash.winfo_screenwidth()
    screen_height = splash.winfo_screenheight()
    x = (screen_width - 400) // 2
    y = (screen_height - 230) // 2
    splash.geometry(f"400x230+{x}+{y}")
    splash.overrideredirect(True)
    
    # Header Title
    lbl_title = tk.Label(splash, text="IFMIS Suite Portal", font=("Segoe UI", 16, "bold"), fg="#00c9a7", bg="#0d1117")
    lbl_title.pack(pady=(25, 5))
    
    lbl_subtitle = tk.Label(splash, text="OFFICE OF THE ACCOUNTANT GENERAL (A&E)\nMadhya Pradesh, Gwalior", font=("Segoe UI", 8, "bold"), fg="#f8fafc", bg="#0d1117")
    lbl_subtitle.pack(pady=5)
    
    # Progress status label
    lbl_status = tk.Label(splash, text="Initializing environment...", font=("Segoe UI", 9), fg="#8b949e", bg="#0d1117")
    lbl_status.pack(pady=(20, 5))
    
    # Styled progress bar
    style = ttk.Style()
    style.theme_use('default')
    style.configure("Teal.Horizontal.TProgressbar", thickness=8, troughcolor='#161b22', bordercolor='#161b22', background='#00c9a7')
    
    progress = ttk.Progressbar(splash, style="Teal.Horizontal.TProgressbar", length=320, mode='determinate')
    progress.pack(pady=5)
    progress['value'] = 10
    splash.update()
    
    # 2. Setup CustomTkinter dark parameters
    import customtkinter as ctk
    ctk.set_appearance_mode("Dark")
    ctk.set_default_color_theme("blue")
    
    progress['value'] = 25
    lbl_status.configure(text="Awaiting operator identity...")
    splash.update()
    
    # Show Custom Operator Login Dialog
    dialog = OperatorLoginDialog(splash)
    splash.wait_window(dialog)
    operator_name = dialog.result
    
    if not operator_name:
        splash.destroy()
        sys.exit()
        
    progress['value'] = 40
    lbl_status.configure(text="Checking license & security...")
    splash.update()
    
    # Run security HWID & kill-switch checks
    check_security(operator_name)
    
    # Helper callback to update progress during engine imports
    def update_splash(text, val):
        lbl_status.configure(text=text)
        progress['value'] = val
        splash.update()
        
    # Load dynamic heavy engines
    load_engines(progress_callback=update_splash)
    
    # Destroy splash and run main app window
    splash.destroy()
    
    # Run application
    app = IFMISSuiteApp(operator_name=operator_name)
    app.mainloop()
